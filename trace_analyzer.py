"""Torch profiler trace analyzer for kernel execution analysis.

Analyzes torch profiler trace files (.trace.json.gz) and provides breakdowns by:
- Stage: prefill vs decode
- Kernel type: attention, MoE, quantization, communication, linear, memory, other
- Layer-level analysis with per-layer kernel sequences
"""

import argparse
import csv
import gzip
import json
import logging
import os
import re
import sys
from dataclasses import dataclass, field
from enum import Enum
from typing import Any, Dict, List, Optional, Tuple

logger = logging.getLogger(__name__)


class Stage(Enum):
    """Execution stage for a kernel."""

    PREFILL = "prefill"
    DECODE = "decode"
    UNKNOWN = "unknown"


class KernelType(Enum):
    """Classification of kernel by operation type."""

    ATTENTION = "attention"
    MOE = "moe"
    QUANTIZATION = "quantization"
    COMMUNICATION = "communication"
    LINEAR = "linear"
    MEMORY = "memory"
    OTHER = "other"


class LayerType(Enum):
    """Type of transformer layer."""

    MLA_FC = "MLA+FC"  # MLA attention + fully connected
    MLA_MOE = "MLA+MoE"  # MLA attention + MoE
    UNKNOWN = "unknown"


@dataclass
class KernelEvent:
    """Single kernel execution event from the trace."""

    name: str
    timestamp_us: float  # microseconds
    duration_us: float  # microseconds
    stage: Stage
    kernel_type: KernelType
    initial_stage: Stage
    grid: Optional[Tuple[int, int, int]] = None
    block: Optional[Tuple[int, int, int]] = None
    simplified_name: str = ""  # Short name for display

    @property
    def duration_ms(self) -> float:
        return self.duration_us / 1000.0


@dataclass
class KernelStats:
    """Aggregated statistics for a group of kernels."""

    count: int = 0
    total_time_us: float = 0.0
    min_time_us: float = float("inf")
    max_time_us: float = 0.0

    def add(self, duration_us: float) -> None:
        self.count += 1
        self.total_time_us += duration_us
        self.min_time_us = min(self.min_time_us, duration_us)
        self.max_time_us = max(self.max_time_us, duration_us)

    @property
    def avg_time_us(self) -> float:
        return self.total_time_us / self.count if self.count > 0 else 0.0

    @property
    def total_time_ms(self) -> float:
        return self.total_time_us / 1000.0

    @property
    def avg_time_ms(self) -> float:
        return self.avg_time_us / 1000.0

    @property
    def min_time_ms(self) -> float:
        return self.min_time_us / 1000.0 if self.min_time_us != float("inf") else 0.0

    @property
    def max_time_ms(self) -> float:
        return self.max_time_us / 1000.0


@dataclass
class LayerEvent:
    """Single layer's execution with all its kernels."""

    layer_idx: int
    layer_type: LayerType
    stage: Stage
    kernels: List[KernelEvent] = field(default_factory=list)

    @property
    def total_time_us(self) -> float:
        return sum(k.duration_us for k in self.kernels)

    @property
    def total_time_ms(self) -> float:
        return self.total_time_us / 1000.0

    @property
    def attention_time_us(self) -> float:
        return sum(
            k.duration_us for k in self.kernels if k.kernel_type == KernelType.ATTENTION
        )

    @property
    def moe_time_us(self) -> float:
        return sum(
            k.duration_us for k in self.kernels if k.kernel_type == KernelType.MOE
        )

    @property
    def linear_time_us(self) -> float:
        return sum(
            k.duration_us for k in self.kernels if k.kernel_type == KernelType.LINEAR
        )

    @property
    def communication_time_us(self) -> float:
        return sum(
            k.duration_us
            for k in self.kernels
            if k.kernel_type == KernelType.COMMUNICATION
        )

    @property
    def quantization_time_us(self) -> float:
        return sum(
            k.duration_us
            for k in self.kernels
            if k.kernel_type == KernelType.QUANTIZATION
        )

    def get_breakdown(self) -> Dict[KernelType, float]:
        """Get time breakdown by kernel type."""
        breakdown = {}
        for k in self.kernels:
            if k.kernel_type not in breakdown:
                breakdown[k.kernel_type] = 0.0
            breakdown[k.kernel_type] += k.duration_us
        return breakdown


@dataclass
class AnalysisResult:
    """Complete analysis result with all breakdowns."""

    # All kernel events (time-ordered)
    events: List[KernelEvent] = field(default_factory=list)

    # Per-kernel stats (keyed by kernel name)
    per_kernel_stats: Dict[str, KernelStats] = field(default_factory=dict)

    # Per-stage stats
    per_stage_stats: Dict[Stage, KernelStats] = field(default_factory=dict)

    # Per-type stats
    per_type_stats: Dict[KernelType, KernelStats] = field(default_factory=dict)

    # Per-stage per-type stats
    per_stage_type_stats: Dict[Tuple[Stage, KernelType], KernelStats] = field(
        default_factory=dict
    )

    # Per-stage per-kernel stats
    per_stage_kernel_stats: Dict[Tuple[Stage, str], KernelStats] = field(
        default_factory=dict
    )

    # Layer-level analysis
    layers: List[LayerEvent] = field(default_factory=list)

    # TTFT/ITL metrics (elapsed wall-clock time, not sum of durations)
    ttft_us: float = 0.0  # Time To First Token (prefill elapsed time)
    itl_us: float = 0.0  # Inter-Token Latency (avg decode iteration time)
    decode_iterations: int = 0  # Number of decode iterations
    prefill_elapsed_us: float = 0.0  # Total prefill elapsed time
    decode_elapsed_us: float = 0.0  # Total decode elapsed time

    # Correction factors for per-request and MTP-adjusted metrics
    num_requests: int = 1  # Number of requests in trace
    accept_length: float = 1.0  # MTP acceptance length

    @property
    def total_time_us(self) -> float:
        return sum(e.duration_us for e in self.events)

    @property
    def total_time_ms(self) -> float:
        return self.total_time_us / 1000.0

    @property
    def total_time_s(self) -> float:
        return self.total_time_us / 1_000_000.0

    @property
    def ttft_ms(self) -> float:
        return self.ttft_us / 1000.0

    @property
    def itl_ms(self) -> float:
        return self.itl_us / 1000.0

    @property
    def ttft_corrected_us(self) -> float:
        """TTFT corrected for number of requests (per-request TTFT)."""
        if self.num_requests > 0:
            return self.ttft_us / self.num_requests
        return self.ttft_us

    @property
    def ttft_corrected_ms(self) -> float:
        return self.ttft_corrected_us / 1000.0

    @property
    def itl_corrected_us(self) -> float:
        """ITL corrected for MTP acceptance length."""
        if self.accept_length > 0:
            return self.itl_us / self.accept_length
        return self.itl_us

    @property
    def itl_corrected_ms(self) -> float:
        return self.itl_corrected_us / 1000.0

    def get_stage_elapsed_us(self, stage: Stage) -> float:
        """Calculate elapsed wall-clock time for a stage using timestamps.

        Elapsed time = (max kernel end timestamp) - (min kernel start timestamp)
        This accounts for kernel parallelism, unlike sum of durations.
        """
        stage_events = [e for e in self.events if e.stage == stage]
        if not stage_events:
            return 0.0
        min_start = min(e.timestamp_us for e in stage_events)
        max_end = max(e.timestamp_us + e.duration_us for e in stage_events)
        return max_end - min_start

    def count_decode_iterations(self) -> int:
        """Count decode iterations using FC layer count.

        Each decode forward pass has exactly one FC layer (MLA+FC type).
        This gives the number of decode iterations (tokens generated).
        """
        if not self.layers:
            return 0
        decode_fc_layers = [
            l for l in self.layers
            if l.stage == Stage.DECODE and l.layer_type == LayerType.MLA_FC
        ]
        return len(decode_fc_layers)


class KernelClassifier:
    """Classifies kernels by stage, type, and provides simplified names."""

    def __init__(self, grid_threshold: int = 10000):
        self.grid_threshold = grid_threshold

        # Stage detection patterns (priority order)
        self._stage_force_unknown_patterns = [
            re.compile(
                r"_fused_rms_fp8|_fused_rms_mxfp4|_gemm_a8w8|_batched_gemm_a8w8|"
                r"_gemm_afp4wfp4|_batched_gemm_a16wfp4",
                re.IGNORECASE,
            ),
        ]
        self._stage_patterns = [
            (re.compile(r"set_mla_kv_buffer_kernel|concat_and_cast_mha_k_kernel", re.IGNORECASE), Stage.PREFILL),
            (re.compile(r"qseqlen[1-4][^0-9]|qseqlen[1-4]$", re.IGNORECASE), Stage.DECODE),
            (re.compile(r"qseqlen[5-9]|qseqlen\d{2,}", re.IGNORECASE), Stage.PREFILL),
            (re.compile(r"_decode_|decode_attention", re.IGNORECASE), Stage.DECODE),
            (re.compile(r"_prefill_|flash_attn|fmha_fwd", re.IGNORECASE), Stage.PREFILL),
            (re.compile(r"generate_draft_decode|create_extend_after_decode"), Stage.DECODE),
        ]

        # Kernel type patterns
        self._type_patterns = [
            # Attention patterns
            (
                re.compile(
                    r"aiter::mla_|mla_a8w8|decode_attention|flash_attn|attention|softmax|"
                    r"fmha_|mla_reduce|kv_cache|flashinfer|set_mla_kv|"
                    r"kn_entry_2c_sbhd|kn_get_mla_metadata|qk_rope|"
                    r"concat_and_cast_mha|kn_mla_reduce",
                    re.IGNORECASE,
                ),
                KernelType.ATTENTION,
            ),
            # MoE patterns
            (
                re.compile(
                    r"fused_moe|moe_align|topk|expert|MoeFlatmm|MoeSorting|"
                    r"kernel_moe_gemm|kernel_moe_mxgemm|shared_experts|grouped_topk|"
                    r"moe_fused_gate|_moe_mxfp4_sort",
                    re.IGNORECASE,
                ),
                KernelType.MOE,
            ),
            # Quantization patterns
            (
                re.compile(
                    r"mxfp4|fp8|quant|_gemm_afp4wfp4|_fused_rms_mxfp4|"
                    r"_batched_gemm_a16wfp4|dynamic_per_group_scaled_quant|"
                    r"_dynamic_mxfp4|_fused_flatten|fp4x2|_gemm_a8w8|"
                    r"_batched_gemm_a8w8|_fused_rms_fp8",
                    re.IGNORECASE,
                ),
                KernelType.QUANTIZATION,
            ),
            # Communication patterns
            (
                re.compile(
                    r"all_reduce|cross_device_reduce|nccl|rccl|broadcast|"
                    r"allgather|reduce_scatter|rcclGenericKernel",
                    re.IGNORECASE,
                ),
                KernelType.COMMUNICATION,
            ),
            # Linear/GEMM patterns (after quant to avoid overlap)
            (
                re.compile(
                    r"Cijk_Alik_Bljk|_gemm_a16_w16|Custom_Cijk",
                    re.IGNORECASE,
                ),
                KernelType.LINEAR,
            ),
            # Memory patterns
            (
                re.compile(
                    r"memcpy|memset|__amd_rocclr_copyBuffer|"
                    r"bfloat16_copy|float8_copy",
                    re.IGNORECASE,
                ),
                KernelType.MEMORY,
            ),
        ]

        # Simplified name patterns for display
        self._simplify_patterns = [
            (re.compile(r"ncclDevKernel|rcclGenericKernel|cross_device_reduce"), "ALLREDUCE"), 
            (re.compile(r"fmha_fwd"), "FMHA"),
            (re.compile(r"mla_a8w8.*qseqlen1"), "MLA_DECODE"),
            (
                re.compile(r"mla_a8w8.*qseqlen[2-9]|mla_a8w8.*qseqlen\d{2,}"),
                "MLA_PREFILL",
            ),
            (re.compile(r"kn_mla_reduce"), "MLA_REDUCE"),
            (re.compile(r"_fused_rms_mxfp4_quant"), "RMS_MXFP4"),
            (re.compile(r"_fused_rms_fp8"), "RMS_FP8"),
            (re.compile(r"_gemm_afp4wfp4.*reduce"), "GEMM_FP4_RED"),
            (re.compile(r"_gemm_afp4wfp4"), "GEMM_FP4"),
            (re.compile(r"_gemm_a8w8_blockscale.*reduce"), "GEMM_FP8_RED"),
            (re.compile(r"_gemm_a8w8_blockscale"), "GEMM_FP8"),
            (re.compile(r"Rmsnorm2dFwd"), "RMSNORM"),
            (re.compile(r"kn_entry_2c_sbhd"), "KV_CACHE"),
            (re.compile(r"set_mla_kv_buffer"), "MLA_KV_SET"),
            (re.compile(r"_dynamic_mxfp4_quant"), "DYN_MXFP4"),
            (re.compile(r"concat_and_cast_mha"), "MHA_CONCAT"),
            (re.compile(r"kernel_moe_mxgemm|kernel_moe_gemm"), "MOE_GEMM"),
            (re.compile(r"MoeSorting|moe_fused_gate|_moe_mxfp4_sort"), "MOE_SORT"),
            (re.compile(r"dynamic_per_group_scaled_quant"), "DYN_QUANT"),
            (re.compile(r"fused_append_shared_experts"), "SHARED_EXP"),
            (re.compile(r"act_and_mul_kernel"), "ACT_MUL"),
            (re.compile(r"Cijk_Alik|Custom_Cijk"), "HIPBLAS_GEMM"),
            (re.compile(r"grouped_topk"), "TOPK"),
            (re.compile(r"bfloat16_copy"), "BF16_COPY"),
            (re.compile(r"float8_copy"), "FP8_COPY"),
            (re.compile(r"create_flashinfer|generate_draft"), "SPEC_INDEX"),
            (re.compile(r"qk_rope_cat_and_cache"), "ROPE_CACHE"),
            (re.compile(r"_batched_gemm_a16wfp4"), "BATCH_GEMM_FP4"),
            (re.compile(r"_batched_gemm_a8w8"), "BATCH_GEMM_FP8"),
            (re.compile(r"_gemm_a16_w16"), "GEMM_A16W16"),
            (re.compile(r"_fused_flatten"), "FLATTEN_QUANT"),
            (re.compile(r"MoeFlatmm"), "MOE_FLATMM"),
        ]

    def classify_stage(
        self, name: str, grid: Optional[Tuple[int, int, int]] = None
    ) -> Stage:
        """Classify kernel stage based on name and grid dimensions."""
        for pattern in self._stage_force_unknown_patterns:
            if pattern.search(name):
                return Stage.UNKNOWN
        for pattern, stage in self._stage_patterns:
            if pattern.search(name):
                return stage

        if grid is not None and len(grid) >= 1:
            grid_0 = grid[0]
            if grid_0 > self.grid_threshold:
                return Stage.DECODE
            elif grid_0 < 200:
                return Stage.PREFILL

        return Stage.UNKNOWN

    def classify_type(self, name: str) -> KernelType:
        """Classify kernel type based on name patterns."""
        for pattern, kernel_type in self._type_patterns:
            if pattern.search(name):
                return kernel_type
        return KernelType.OTHER

    def simplify_name(self, name: str) -> str:
        """Get simplified kernel name for display."""
        for pattern, simple_name in self._simplify_patterns:
            if pattern.search(name):
                return simple_name
        # Truncate unknown names
        return name[:25] if len(name) > 25 else name


class LayerDetector:
    """Detects layer boundaries in kernel sequences."""

    def __init__(self, mtp_qseqlen_decode: bool = False):
        self._mtp_qseqlen_decode = mtp_qseqlen_decode
        # Layer start markers (after ALLREDUCE)
        self._layer_start_patterns = [
            re.compile(
                r"_fused_rms_mxfp4_quant|_fused_rms_fp8"
            ),  # Prefill/decode layernorm
        ]

        # MoE block markers
        self._moe_markers = [
            re.compile(
                r"MoeSorting|moe_fused_gate|kernel_moe_gemm|kernel_moe_mxgemm|_moe_mxfp4_sort"
            ),
        ]

        # FC block markers (activation function after linear)
        self._fc_markers = [
            re.compile(r"act_and_mul_kernel|silu_kernel"),
        ]

        # Attention stage hints for layer stage inference
        self._attention_decode_patterns = [
            re.compile(r"qseqlen1[^0-9]|qseqlen1$|decode_attention", re.IGNORECASE),
            re.compile(r"mla_a8w8.*qseqlen(1|4)", re.IGNORECASE),
        ]
        if self._mtp_qseqlen_decode:
            self._attention_decode_patterns.append(
                re.compile(r"qseqlen[2-9]|qseqlen\d{2,}", re.IGNORECASE)
            )
        self._attention_prefill_patterns = [
            re.compile(r"fmha_fwd|flash_attn", re.IGNORECASE),
        ]
        if not self._mtp_qseqlen_decode:
            self._attention_prefill_patterns.append(
                re.compile(r"qseqlen[2-9]|qseqlen\d{2,}", re.IGNORECASE)
            )

    def detect_layers(self, events: List[KernelEvent]) -> List[LayerEvent]:
        """Detect layer boundaries and classify layers.

        Layer stage is determined by the dominant stage of kernels within each layer.
        """
        if not events:
            return []

        layers = []
        current_layer_kernels: List[KernelEvent] = []
        layer_idx = 0
        in_layer = False
        saw_attention = False
        saw_moe = False
        saw_fc = False

        for i, event in enumerate(events):
            name = event.name
            simplified = event.simplified_name

            # Detect layer start: ALLREDUCE followed by RMS layernorm
            is_allreduce = (
                simplified == "ALLREDUCE"
                or "rcclGenericKernel" in name
                or "cross_device_reduce" in name
            )
            is_layer_start = any(p.search(name) for p in self._layer_start_patterns)

            # Check if this is a MoE or FC marker
            is_moe = any(p.search(name) for p in self._moe_markers)
            is_fc = any(p.search(name) for p in self._fc_markers)
            is_attention = event.kernel_type == KernelType.ATTENTION

            # Start a new layer on layer_start after allreduce or at very beginning
            if is_layer_start and (
                not in_layer
                or (
                    len(current_layer_kernels) > 0
                    and current_layer_kernels[-1].simplified_name == "ALLREDUCE"
                )
            ):
                # Save previous layer if exists
                if current_layer_kernels and saw_attention:
                    layer_type = (
                        LayerType.MLA_MOE
                        if saw_moe
                        else (LayerType.MLA_FC if saw_fc else LayerType.UNKNOWN)
                    )
                    # Determine stage from kernels
                    stage = self._determine_layer_stage(current_layer_kernels)
                    layers.append(
                        LayerEvent(
                            layer_idx=layer_idx,
                            layer_type=layer_type,
                            stage=stage,
                            kernels=current_layer_kernels.copy(),
                        )
                    )
                    layer_idx += 1

                # Start new layer (include the preceding ALLREDUCE if present)
                if (
                    current_layer_kernels
                    and current_layer_kernels[-1].simplified_name == "ALLREDUCE"
                ):
                    current_layer_kernels = [current_layer_kernels[-1], event]
                else:
                    current_layer_kernels = [event]
                in_layer = True
                saw_attention = False
                saw_moe = False
                saw_fc = False
            else:
                current_layer_kernels.append(event)

            # Track what we've seen in this layer
            if is_attention:
                saw_attention = True
            if is_moe:
                saw_moe = True
            if is_fc:
                saw_fc = True

        # Save last layer
        if current_layer_kernels and saw_attention:
            layer_type = (
                LayerType.MLA_MOE
                if saw_moe
                else (LayerType.MLA_FC if saw_fc else LayerType.UNKNOWN)
            )
            stage = self._determine_layer_stage(current_layer_kernels)
            layers.append(
                LayerEvent(
                    layer_idx=layer_idx,
                    layer_type=layer_type,
                    stage=stage,
                    kernels=current_layer_kernels,
                )
            )

        return layers

    def _determine_layer_stage(self, kernels: List[KernelEvent]) -> Stage:
        """Determine layer stage based on dominant stage of its kernels."""
        prefill_count = sum(1 for k in kernels if k.stage == Stage.PREFILL)
        decode_count = sum(1 for k in kernels if k.stage == Stage.DECODE)

        if prefill_count > decode_count:
            return Stage.PREFILL
        if decode_count > prefill_count:
            return Stage.DECODE

        return Stage.UNKNOWN

    def _infer_stage_from_names(self, kernels: List[KernelEvent]) -> Stage:
        for k in kernels:
            name = k.name
            if any(p.search(name) for p in self._attention_decode_patterns):
                return Stage.DECODE
            if any(p.search(name) for p in self._attention_prefill_patterns):
                return Stage.PREFILL
        return Stage.UNKNOWN


class TraceAnalyzer:
    """Main analyzer that loads and processes trace files."""

    def __init__(self, grid_threshold: int = 10000, mtp_qseqlen_decode: bool = False):
        self.classifier = KernelClassifier(grid_threshold=grid_threshold)
        self.layer_detector = LayerDetector(mtp_qseqlen_decode=mtp_qseqlen_decode)

    def load_trace(self, path: str) -> Dict[str, Any]:
        """Load trace file (supports .json.gz and .json)."""
        logger.info(f"Loading trace file: {path}")

        if path.endswith(".gz"):
            with gzip.open(path, "rt", encoding="utf-8") as f:
                return json.load(f)
        else:
            with open(path, "r", encoding="utf-8") as f:
                return json.load(f)

    def _parse_grid(self, grid_data: Any) -> Optional[Tuple[int, int, int]]:
        """Parse grid dimensions from trace data."""
        if grid_data is None:
            return None

        if isinstance(grid_data, list) and len(grid_data) >= 3:
            return (int(grid_data[0]), int(grid_data[1]), int(grid_data[2]))

        if isinstance(grid_data, str):
            try:
                parsed = json.loads(grid_data)
                if isinstance(parsed, list) and len(parsed) >= 3:
                    return (int(parsed[0]), int(parsed[1]), int(parsed[2]))
            except (json.JSONDecodeError, ValueError):
                pass

        return None

    def analyze(
        self, trace_data: Dict[str, Any], detect_layers: bool = False
    ) -> AnalysisResult:
        """Analyze trace data and compute statistics."""
        result = AnalysisResult()

        events = trace_data.get("traceEvents", [])
        logger.info(f"Total events in trace: {len(events)}")

        # Filter kernel events
        kernel_events = [
            e for e in events if e.get("ph") == "X" and e.get("cat") == "kernel"
        ]
        logger.info(f"Kernel events found: {len(kernel_events)}")

        # Sort by timestamp
        kernel_events.sort(key=lambda x: x.get("ts", 0))

        # Process each kernel event
        for event in kernel_events:
            name = event.get("name", "")
            timestamp_us = event.get("ts", 0)
            duration_us = event.get("dur", 0)

            args = event.get("args", {})
            grid = self._parse_grid(args.get("grid"))
            block = self._parse_grid(args.get("block"))

            # Classify
            stage = self.classifier.classify_stage(name, grid)
            kernel_type = self.classifier.classify_type(name)
            simplified_name = self.classifier.simplify_name(name)

            if kernel_type in (
                KernelType.COMMUNICATION,
                KernelType.QUANTIZATION,
                KernelType.MOE,
                KernelType.LINEAR,
                KernelType.MEMORY,
            ):
                stage = Stage.UNKNOWN

            # Create event
            kernel_event = KernelEvent(
                name=name,
                timestamp_us=timestamp_us,
                duration_us=duration_us,
                stage=stage,
                kernel_type=kernel_type,
                initial_stage=stage,
                grid=grid,
                block=block,
                simplified_name=simplified_name,
            )
            result.events.append(kernel_event)

            # Update stats
            if name not in result.per_kernel_stats:
                result.per_kernel_stats[name] = KernelStats()
            result.per_kernel_stats[name].add(duration_us)

            if stage not in result.per_stage_stats:
                result.per_stage_stats[stage] = KernelStats()
            result.per_stage_stats[stage].add(duration_us)

            if kernel_type not in result.per_type_stats:
                result.per_type_stats[kernel_type] = KernelStats()
            result.per_type_stats[kernel_type].add(duration_us)

            stage_type_key = (stage, kernel_type)
            if stage_type_key not in result.per_stage_type_stats:
                result.per_stage_type_stats[stage_type_key] = KernelStats()
            result.per_stage_type_stats[stage_type_key].add(duration_us)

            stage_kernel_key = (stage, name)
            if stage_kernel_key not in result.per_stage_kernel_stats:
                result.per_stage_kernel_stats[stage_kernel_key] = KernelStats()
            result.per_stage_kernel_stats[stage_kernel_key].add(duration_us)

        # Detect layers if requested
        if detect_layers:
            # Detect layers from ALL events (don't filter by stage)
            # Layer stage is determined by dominant stage within each layer
            all_layers = self.layer_detector.detect_layers(result.events)

            # Count layers by stage
            prefill_count = sum(1 for l in all_layers if l.stage == Stage.PREFILL)
            decode_count = sum(1 for l in all_layers if l.stage == Stage.DECODE)

            result.layers = all_layers
            logger.info(
                f"Detected {prefill_count} prefill layers, {decode_count} decode layers"
            )
            # Re-assign kernel stages based on layer vote (only for prefill/decode layers)
            for layer in result.layers:
                if layer.stage in (Stage.PREFILL, Stage.DECODE):
                    for k in layer.kernels:
                        k.stage = layer.stage

            # Recompute stage-related stats after layer assignment
            result.per_stage_stats = {}
            result.per_stage_type_stats = {}
            result.per_stage_kernel_stats = {}
            for event in result.events:
                stage = event.stage
                kernel_type = event.kernel_type
                name = event.name
                duration_us = event.duration_us

                if stage not in result.per_stage_stats:
                    result.per_stage_stats[stage] = KernelStats()
                result.per_stage_stats[stage].add(duration_us)

                stage_type_key = (stage, kernel_type)
                if stage_type_key not in result.per_stage_type_stats:
                    result.per_stage_type_stats[stage_type_key] = KernelStats()
                result.per_stage_type_stats[stage_type_key].add(duration_us)

                stage_kernel_key = (stage, name)
                if stage_kernel_key not in result.per_stage_kernel_stats:
                    result.per_stage_kernel_stats[stage_kernel_key] = KernelStats()
                result.per_stage_kernel_stats[stage_kernel_key].add(duration_us)

        # Calculate TTFT and ITL metrics using elapsed time (wall-clock), not sum of durations
        # TTFT = elapsed time for prefill phase
        # ITL = elapsed time for decode phase / number of decode iterations
        result.prefill_elapsed_us = result.get_stage_elapsed_us(Stage.PREFILL)
        result.decode_elapsed_us = result.get_stage_elapsed_us(Stage.DECODE)
        result.ttft_us = result.prefill_elapsed_us

        # Count decode iterations using FC layer count (each iteration has 1 FC layer)
        if detect_layers:
            result.decode_iterations = result.count_decode_iterations()
        else:
            # Fallback: estimate from layer count if layers not detected
            # This is less accurate but provides an estimate
            result.decode_iterations = 0

        # Calculate ITL
        if result.decode_iterations > 0:
            result.itl_us = result.decode_elapsed_us / result.decode_iterations
        else:
            result.itl_us = 0.0

        logger.info(
            f"TTFT: {result.ttft_us/1000:.2f} ms, "
            f"ITL: {result.itl_us/1000:.2f} ms ({result.decode_iterations} decode iterations)"
        )

        return result


class ReportFormatter:
    """Formats analysis results for console output."""

    def __init__(self, result: AnalysisResult):
        self.result = result

    def format_time(self, time_ms: float) -> str:
        """Format time value with appropriate unit."""
        if time_ms >= 1000:
            return f"{time_ms / 1000:.3f} s"
        elif time_ms >= 1:
            return f"{time_ms:.3f} ms"
        else:
            return f"{time_ms * 1000:.3f} us"

    def format_time_us(self, time_us: float) -> str:
        """Format time in microseconds with appropriate unit."""
        return self.format_time(time_us / 1000.0)

    def format_percentage(self, part: float, total: float) -> str:
        """Format percentage."""
        if total == 0:
            return "0.0%"
        return f"{100.0 * part / total:.1f}%"

    def print_summary(self) -> None:
        """Print overall summary."""
        print("=" * 80)
        print("TRACE ANALYSIS SUMMARY")
        print("=" * 80)
        print()

        total_time_ms = self.result.total_time_ms
        print(f"Total Kernel Time:   {self.format_time(total_time_ms)}")

        for stage in [Stage.PREFILL, Stage.DECODE, Stage.UNKNOWN]:
            stats = self.result.per_stage_stats.get(stage, KernelStats())
            stage_time_ms = stats.total_time_ms
            pct = self.format_percentage(stage_time_ms, total_time_ms)
            print(
                f"  - {stage.value.capitalize():12s} {self.format_time(stage_time_ms):>12s} ({pct})"
            )

        print()

        # TTFT and ITL metrics (wall-clock elapsed time)
        print("-" * 80)
        print("TTFT / ITL METRICS (Wall-Clock Time)")
        print("-" * 80)
        print()
        print("Calculation method:")
        print("  TTFT = (last prefill kernel end timestamp) - (first prefill kernel start timestamp)")
        print("  ITL  = (decode elapsed time) / (number of decode iterations)")
        print("  Decode iterations = count of MLA+FC layers in decode stage (1 per forward pass)")
        if self.result.num_requests > 1:
            print(f"  TTFT_corrected = TTFT / num_requests ({self.result.num_requests})")
        if self.result.accept_length != 1.0:
            print(f"  ITL_corrected = ITL / accept_length ({self.result.accept_length})")
        print()
        print(f"{'Metric':<28} {'Raw':>14} {'Corrected':>14} {'Description'}")
        print("-" * 85)
        # TTFT
        ttft_raw = self.format_time(self.result.ttft_ms)
        ttft_corr = self.format_time(self.result.ttft_corrected_ms) if self.result.num_requests > 1 else "-"
        print(f"{'TTFT (prefill time)':<28} {ttft_raw:>14} {ttft_corr:>14} Time to first token")
        # Prefill elapsed
        print(f"{'Prefill elapsed':<28} {self.format_time(self.result.prefill_elapsed_us/1000):>14} {'-':>14} Wall-clock prefill duration")
        # Decode elapsed
        print(f"{'Decode elapsed':<28} {self.format_time(self.result.decode_elapsed_us/1000):>14} {'-':>14} Wall-clock decode duration")
        # Decode iterations
        print(f"{'Decode iterations':<28} {self.result.decode_iterations:>14} {'-':>14} Forward passes (FC layers)")
        # ITL
        itl_raw = self.format_time(self.result.itl_ms)
        itl_corr = self.format_time(self.result.itl_corrected_ms) if self.result.accept_length != 1.0 else "-"
        print(f"{'ITL (per token)':<28} {itl_raw:>14} {itl_corr:>14} Inter-token latency")
        print()
        # Show correction parameters if applied
        if self.result.num_requests > 1 or self.result.accept_length != 1.0:
            print("Correction parameters:")
            if self.result.num_requests > 1:
                print(f"  --num-requests {self.result.num_requests}")
            if self.result.accept_length != 1.0:
                print(f"  --accept-length {self.result.accept_length}")
            print()

    def print_type_breakdown(self) -> None:
        """Print breakdown by kernel type."""
        print("-" * 80)
        print("BREAKDOWN BY KERNEL TYPE")
        print("-" * 80)
        print()
        print(f"{'Type':<16} {'Count':>10} {'Total':>14} {'Avg':>12} {'%':>8}")
        print("-" * 64)

        total_time_ms = self.result.total_time_ms

        sorted_types = sorted(
            self.result.per_type_stats.items(),
            key=lambda x: x[1].total_time_us,
            reverse=True,
        )

        for kernel_type, stats in sorted_types:
            pct = self.format_percentage(stats.total_time_ms, total_time_ms)
            print(
                f"{kernel_type.value:<16} {stats.count:>10,} {self.format_time(stats.total_time_ms):>14} "
                f"{self.format_time(stats.avg_time_ms):>12} {pct:>8}"
            )

        print()

    def print_stage_type_breakdown(self, stage: Stage) -> None:
        """Print type breakdown for a specific stage."""
        print("-" * 80)
        print(f"BREAKDOWN BY TYPE ({stage.value.upper()})")
        print("-" * 80)
        print()
        print(f"{'Type':<16} {'Count':>10} {'Total':>14} {'Avg':>12} {'%':>8}")
        print("-" * 64)

        stage_stats = self.result.per_stage_stats.get(stage, KernelStats())
        stage_total_ms = stage_stats.total_time_ms

        stage_type_items = [
            (kt, stats)
            for (s, kt), stats in self.result.per_stage_type_stats.items()
            if s == stage
        ]
        sorted_items = sorted(
            stage_type_items, key=lambda x: x[1].total_time_us, reverse=True
        )

        for kernel_type, stats in sorted_items:
            pct = self.format_percentage(stats.total_time_ms, stage_total_ms)
            print(
                f"{kernel_type.value:<16} {stats.count:>10,} {self.format_time(stats.total_time_ms):>14} "
                f"{self.format_time(stats.avg_time_ms):>12} {pct:>8}"
            )

        print()

    def print_top_kernels(self, stage: Optional[Stage] = None, top_n: int = 20) -> None:
        """Print top kernels by total time."""
        if stage is not None:
            title = f"TOP {top_n} KERNELS ({stage.value.upper()})"
            kernel_items = [
                (name, stats)
                for (s, name), stats in self.result.per_stage_kernel_stats.items()
                if s == stage
            ]
            stage_stats = self.result.per_stage_stats.get(stage, KernelStats())
            total_time_ms = stage_stats.total_time_ms
        else:
            title = f"TOP {top_n} KERNELS (ALL)"
            kernel_items = list(self.result.per_kernel_stats.items())
            total_time_ms = self.result.total_time_ms

        print("-" * 80)
        print(title)
        print("-" * 80)
        print()
        print(
            f"{'#':<4} {'Count':>10} {'Total':>14} {'Avg':>12} {'%':>8}  {'Kernel Name'}"
        )
        print("-" * 80)

        sorted_kernels = sorted(
            kernel_items, key=lambda x: x[1].total_time_us, reverse=True
        )

        for i, (name, stats) in enumerate(sorted_kernels[:top_n], 1):
            pct = self.format_percentage(stats.total_time_ms, total_time_ms)
            display_name = name[:60] + "..." if len(name) > 63 else name
            print(
                f"{i:<4} {stats.count:>10,} {self.format_time(stats.total_time_ms):>14} "
                f"{self.format_time(stats.avg_time_ms):>12} {pct:>8}  {display_name}"
            )

        print()

    def print_layer_summary(self, stage: Optional[Stage] = None) -> None:
        """Print layer-level summary."""
        if not self.result.layers:
            print("No layer analysis available. Use --layer-analysis to enable.")
            return

        # Filter layers by stage if specified
        layers = self.result.layers
        if stage is not None:
            layers = [l for l in layers if l.stage == stage]

        if not layers:
            return

        stage_name = stage.value.upper() if stage else "ALL"
        print("-" * 100)
        print(f"LAYER ANALYSIS ({stage_name})")
        print("-" * 100)
        print()
        print(
            f"{'Layer':<8} {'Type':<10} {'Total':>12} {'Attn':>12} {'MoE/FC':>12} {'Comm':>12} {'Quant':>12} {'#Kern':>8}"
        )
        print("-" * 100)

        for layer in layers:
            moe_fc_time = (
                layer.moe_time_us
                if layer.layer_type == LayerType.MLA_MOE
                else layer.linear_time_us
            )
            print(
                f"{layer.layer_idx:<8} {layer.layer_type.value:<10} "
                f"{self.format_time_us(layer.total_time_us):>12} "
                f"{self.format_time_us(layer.attention_time_us):>12} "
                f"{self.format_time_us(moe_fc_time):>12} "
                f"{self.format_time_us(layer.communication_time_us):>12} "
                f"{self.format_time_us(layer.quantization_time_us):>12} "
                f"{len(layer.kernels):>8}"
            )

        print()

    def print_layer_detail(self, layer_idx: int, stage: Optional[Stage] = None) -> None:
        """Print detailed kernel sequence for a specific layer."""
        if not self.result.layers:
            print("No layer analysis available. Use --layer-analysis to enable.")
            return

        # Find the layer
        layer = None
        for l in self.result.layers:
            if l.layer_idx == layer_idx and (stage is None or l.stage == stage):
                layer = l
                break

        if layer is None:
            print(f"Layer {layer_idx} not found.")
            return

        # Filter out trailing ALLREDUCE (belongs to next layer)
        kernels = layer.kernels
        if kernels and kernels[-1].simplified_name == "ALLREDUCE":
            kernels = kernels[:-1]

        # Recalculate total time without trailing ALLREDUCE
        total_time = sum(k.duration_us for k in kernels)

        print("-" * 180)
        print(
            f"LAYER {layer.layer_idx} DETAIL ({layer.layer_type.value}, {layer.stage.value.upper()})"
        )
        print("-" * 180)
        print()
        print(
            f"{'#':<4} {'Duration (us)':>14} {'%':>7} {'Type':<14} {'Short Name':<18} {'Kernel Name'}"
        )
        print("-" * 180)

        for i, kernel in enumerate(kernels, 1):
            pct = self.format_percentage(kernel.duration_us, total_time)
            short_name = kernel.simplified_name if kernel.simplified_name else "-"
            # Truncate kernel name to 50 characters
            display_name = kernel.name[:50] + "..." if len(kernel.name) > 50 else kernel.name
            print(
                f"{i:<4} {kernel.duration_us:>14.3f} {pct:>7} "
                f"{kernel.kernel_type.value:<14} {short_name:<18} {display_name}"
            )

        print()
        print(f"Layer Total: {total_time:.3f} us ({total_time/1000:.3f} ms)")

        # Print breakdown (recalculated without trailing ALLREDUCE)
        breakdown = {}
        for k in kernels:
            if k.kernel_type not in breakdown:
                breakdown[k.kernel_type] = 0.0
            breakdown[k.kernel_type] += k.duration_us

        print()
        print("Breakdown:")
        for kt, time_us in sorted(breakdown.items(), key=lambda x: x[1], reverse=True):
            pct = self.format_percentage(time_us, total_time)
            print(f"  - {kt.value:<14} {time_us:>14.3f} us ({pct})")

        print()

    def print_full_report(
        self, top_n: int = 20, stage_filter: Optional[str] = None
    ) -> None:
        """Print full analysis report."""
        self.print_summary()
        self.print_type_breakdown()

        if stage_filter is None or stage_filter == "all":
            for stage in [Stage.PREFILL, Stage.DECODE]:
                if stage in self.result.per_stage_stats:
                    self.print_stage_type_breakdown(stage)
                    self.print_top_kernels(stage=stage, top_n=top_n)
        elif stage_filter == "prefill":
            self.print_stage_type_breakdown(Stage.PREFILL)
            self.print_top_kernels(stage=Stage.PREFILL, top_n=top_n)
        elif stage_filter == "decode":
            self.print_stage_type_breakdown(Stage.DECODE)
            self.print_top_kernels(stage=Stage.DECODE, top_n=top_n)

    def print_layer_debug(self, layer_idx: int, top_n: int = 20) -> None:
        """Print debug details for a specific layer's stage decision."""
        if not self.result.layers:
            print("No layer analysis available. Use --layer-analysis to enable.")
            return

        layer = None
        for l in self.result.layers:
            if l.layer_idx == layer_idx:
                layer = l
                break

        if layer is None:
            print(f"Layer {layer_idx} not found.")
            return

        prefill_count = sum(1 for k in layer.kernels if k.initial_stage == Stage.PREFILL)
        decode_count = sum(1 for k in layer.kernels if k.initial_stage == Stage.DECODE)
        unknown_count = sum(1 for k in layer.kernels if k.initial_stage == Stage.UNKNOWN)

        print("-" * 120)
        print(f"LAYER DEBUG {layer.layer_idx} ({layer.layer_type.value})")
        print("-" * 120)
        print(f"Layer stage (after vote): {layer.stage.value}")
        print(
            f"Vote counts (initial stages): prefill={prefill_count}, decode={decode_count}, unknown={unknown_count}"
        )
        print()

        sorted_kernels = sorted(layer.kernels, key=lambda k: k.duration_us, reverse=True)
        print(
            f"{'#':<4} {'Duration (us)':>14} {'Type':<14} {'Init':<8} {'Final':<8} {'Kernel Name'}"
        )
        print("-" * 120)
        for i, k in enumerate(sorted_kernels[:top_n], 1):
            display_name = k.name[:80] + "..." if len(k.name) > 83 else k.name
            print(
                f"{i:<4} {k.duration_us:>14.3f} {k.kernel_type.value:<14} "
                f"{k.initial_stage.value:<8} {k.stage.value:<8} {display_name}"
            )
        print()


class CSVExporter:
    """Exports analysis results to CSV files."""

    def __init__(self, result: AnalysisResult):
        self.result = result

    def export_kernel_stats(self, path: str) -> None:
        """Export per-kernel statistics to CSV."""
        with open(path, "w", newline="") as f:
            writer = csv.writer(f)
            writer.writerow(
                [
                    "kernel_name",
                    "count",
                    "total_time_us",
                    "avg_time_us",
                    "min_time_us",
                    "max_time_us",
                ]
            )

            for name, stats in sorted(
                self.result.per_kernel_stats.items(),
                key=lambda x: x[1].total_time_us,
                reverse=True,
            ):
                writer.writerow(
                    [
                        name,
                        stats.count,
                        f"{stats.total_time_us:.3f}",
                        f"{stats.avg_time_us:.3f}",
                        (
                            f"{stats.min_time_us:.3f}"
                            if stats.min_time_us != float("inf")
                            else "0"
                        ),
                        f"{stats.max_time_us:.3f}",
                    ]
                )

        logger.info(f"Kernel stats exported to: {path}")

    def export_events(self, path: str) -> None:
        """Export all kernel events to CSV."""
        with open(path, "w", newline="") as f:
            writer = csv.writer(f)
            writer.writerow(
                [
                    "timestamp_us",
                    "duration_us",
                    "stage",
                    "kernel_type",
                    "simplified_name",
                    "grid",
                    "block",
                    "kernel_name",
                ]
            )

            for event in self.result.events:
                writer.writerow(
                    [
                        f"{event.timestamp_us:.3f}",
                        f"{event.duration_us:.3f}",
                        event.stage.value,
                        event.kernel_type.value,
                        event.simplified_name,
                        str(event.grid) if event.grid else "",
                        str(event.block) if event.block else "",
                        event.name,
                    ]
                )

        logger.info(f"Events exported to: {path}")

    def export_layers(self, path: str) -> None:
        """Export layer-level statistics to CSV."""
        with open(path, "w", newline="") as f:
            writer = csv.writer(f)
            writer.writerow(
                [
                    "layer_idx",
                    "layer_type",
                    "stage",
                    "total_time_us",
                    "attention_time_us",
                    "moe_time_us",
                    "linear_time_us",
                    "communication_time_us",
                    "quantization_time_us",
                    "kernel_count",
                ]
            )

            for layer in self.result.layers:
                writer.writerow(
                    [
                        layer.layer_idx,
                        layer.layer_type.value,
                        layer.stage.value,
                        f"{layer.total_time_us:.3f}",
                        f"{layer.attention_time_us:.3f}",
                        f"{layer.moe_time_us:.3f}",
                        f"{layer.linear_time_us:.3f}",
                        f"{layer.communication_time_us:.3f}",
                        f"{layer.quantization_time_us:.3f}",
                        len(layer.kernels),
                    ]
                )

        logger.info(f"Layer stats exported to: {path}")

    def export_full_analysis(
        self,
        output_path: str,
        layer_indices: Optional[List[int]] = None,
        stage_filter: Optional[Stage] = None,
    ) -> str:
        """Export full analysis to Excel file with multiple sheets.

        Creates sheets:
        - Summary: Overall summary and breakdown
        - Layer_N: Detailed kernel list for each specified layer

        Args:
            output_path: Output Excel file path (.xlsx)
            layer_indices: List of layer indices to export (None = all layers)
            stage_filter: Filter layers by stage (None = all stages)

        Returns:
            Path to created Excel file
        """
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Alignment, Font, PatternFill
        except ImportError:
            raise ImportError(
                "openpyxl is required for Excel export. Install with: pip install openpyxl"
            )

        wb = Workbook()

        # Create summary sheet
        ws_summary = wb.active
        ws_summary.title = "Summary"
        self._write_summary_sheet(ws_summary)

        # Filter layers
        layers = self.result.layers
        if stage_filter is not None:
            layers = [l for l in layers if l.stage == stage_filter]

        if layer_indices is not None:
            layers = [l for l in layers if l.layer_idx in layer_indices]

        # Create sheet for each layer
        for layer in layers:
            ws_layer = wb.create_sheet(title=f"Layer_{layer.layer_idx}")
            self._write_layer_sheet(ws_layer, layer)

        # Save workbook
        wb.save(output_path)
        logger.info(f"Exported Excel file: {output_path}")
        return output_path

    def _write_summary_sheet(self, ws) -> None:
        """Write summary statistics to Excel sheet."""
        from openpyxl.styles import Alignment, Font, PatternFill

        # Styling
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")

        row = 1

        # Overall summary section
        ws.cell(row=row, column=1, value="=== OVERALL SUMMARY ===").font = Font(bold=True, size=12)
        row += 1

        headers = ["Metric", "Value (us)", "Percentage"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
        row += 1

        total_time_us = self.result.total_time_us
        ws.cell(row=row, column=1, value="Total Kernel Time")
        ws.cell(row=row, column=2, value=round(total_time_us, 3))
        ws.cell(row=row, column=3, value="100.0%")
        row += 1

        for stage in [Stage.PREFILL, Stage.DECODE, Stage.UNKNOWN]:
            stats = self.result.per_stage_stats.get(stage, KernelStats())
            pct = f"{100.0 * stats.total_time_us / total_time_us:.1f}%" if total_time_us > 0 else "0.0%"
            ws.cell(row=row, column=1, value=f"{stage.value.capitalize()} Time")
            ws.cell(row=row, column=2, value=round(stats.total_time_us, 3))
            ws.cell(row=row, column=3, value=pct)
            row += 1

        row += 1

        # TTFT/ITL section
        ws.cell(row=row, column=1, value="=== TTFT / ITL METRICS (Wall-Clock Time) ===").font = Font(bold=True, size=12)
        row += 1

        headers = ["Metric", "Raw (ms)", "Corrected (ms)", "Description"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
        row += 1

        # TTFT
        ws.cell(row=row, column=1, value="TTFT")
        ws.cell(row=row, column=2, value=round(self.result.ttft_us / 1000, 2))
        if self.result.num_requests > 1:
            ws.cell(row=row, column=3, value=round(self.result.ttft_corrected_us / 1000, 2))
        else:
            ws.cell(row=row, column=3, value="-")
        ws.cell(row=row, column=4, value="Time To First Token (prefill elapsed time)")
        row += 1

        # ITL
        ws.cell(row=row, column=1, value="ITL")
        ws.cell(row=row, column=2, value=round(self.result.itl_us / 1000, 2))
        if self.result.accept_length != 1.0:
            ws.cell(row=row, column=3, value=round(self.result.itl_corrected_us / 1000, 2))
        else:
            ws.cell(row=row, column=3, value="-")
        ws.cell(row=row, column=4, value="Inter-Token Latency (avg decode iteration time)")
        row += 1

        # Decode iterations
        ws.cell(row=row, column=1, value="Decode Iterations")
        ws.cell(row=row, column=2, value=self.result.decode_iterations)
        ws.cell(row=row, column=3, value="-")
        ws.cell(row=row, column=4, value="Forward passes (MLA+FC layer count)")
        row += 1

        # Prefill elapsed
        ws.cell(row=row, column=1, value="Prefill Elapsed")
        ws.cell(row=row, column=2, value=round(self.result.prefill_elapsed_us / 1000, 2))
        ws.cell(row=row, column=3, value="-")
        ws.cell(row=row, column=4, value="Wall-clock prefill duration")
        row += 1

        # Decode elapsed
        ws.cell(row=row, column=1, value="Decode Elapsed")
        ws.cell(row=row, column=2, value=round(self.result.decode_elapsed_us / 1000, 2))
        ws.cell(row=row, column=3, value="-")
        ws.cell(row=row, column=4, value="Wall-clock decode duration")
        row += 1

        row += 1

        # Correction parameters
        ws.cell(row=row, column=1, value="Correction Parameters:").font = Font(italic=True)
        row += 1
        ws.cell(row=row, column=1, value=f"num_requests = {self.result.num_requests}")
        ws.cell(row=row, column=2, value="TTFT_corrected = TTFT / num_requests")
        row += 1
        ws.cell(row=row, column=1, value=f"accept_length = {self.result.accept_length}")
        ws.cell(row=row, column=2, value="ITL_corrected = ITL / accept_length")
        row += 1

        row += 1

        # Calculation method note
        ws.cell(row=row, column=1, value="Calculation Method:").font = Font(italic=True)
        row += 1
        ws.cell(row=row, column=1, value="TTFT = (last prefill kernel end timestamp) - (first prefill kernel start timestamp)")
        row += 1
        ws.cell(row=row, column=1, value="ITL = (decode elapsed time) / (decode iterations)")
        row += 1
        ws.cell(row=row, column=1, value="Decode iterations = count of MLA+FC layers in decode stage (1 per forward pass)")
        row += 1

        row += 1

        # Type breakdown section
        ws.cell(row=row, column=1, value="=== BREAKDOWN BY TYPE ===").font = Font(bold=True, size=12)
        row += 1

        headers = ["Type", "Count", "Total (us)", "Avg (us)", "Percentage"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
        row += 1

        sorted_types = sorted(
            self.result.per_type_stats.items(),
            key=lambda x: x[1].total_time_us,
            reverse=True,
        )
        for kernel_type, stats in sorted_types:
            pct = f"{100.0 * stats.total_time_us / total_time_us:.1f}%" if total_time_us > 0 else "0.0%"
            ws.cell(row=row, column=1, value=kernel_type.value)
            ws.cell(row=row, column=2, value=stats.count)
            ws.cell(row=row, column=3, value=round(stats.total_time_us, 3))
            ws.cell(row=row, column=4, value=round(stats.avg_time_us, 3))
            ws.cell(row=row, column=5, value=pct)
            row += 1

        row += 1

        # Layer summary section
        if self.result.layers:
            ws.cell(row=row, column=1, value="=== LAYER SUMMARY ===").font = Font(bold=True, size=12)
            row += 1

            headers = ["Layer", "Type", "Stage", "Total (us)", "Attention (us)", "MoE (us)",
                       "Linear (us)", "Comm (us)", "Quant (us)", "Kernels"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
            row += 1

            for layer in self.result.layers:
                ws.cell(row=row, column=1, value=layer.layer_idx)
                ws.cell(row=row, column=2, value=layer.layer_type.value)
                ws.cell(row=row, column=3, value=layer.stage.value)
                ws.cell(row=row, column=4, value=round(layer.total_time_us, 3))
                ws.cell(row=row, column=5, value=round(layer.attention_time_us, 3))
                ws.cell(row=row, column=6, value=round(layer.moe_time_us, 3))
                ws.cell(row=row, column=7, value=round(layer.linear_time_us, 3))
                ws.cell(row=row, column=8, value=round(layer.communication_time_us, 3))
                ws.cell(row=row, column=9, value=round(layer.quantization_time_us, 3))
                ws.cell(row=row, column=10, value=len(layer.kernels))
                row += 1

        # Auto-adjust column widths
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            ws.column_dimensions[column].width = min(max_length + 2, 50)

    def _write_layer_sheet(self, ws, layer: LayerEvent) -> None:
        """Write layer detail to Excel sheet."""
        from openpyxl.styles import Alignment, Font, PatternFill

        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")

        # Filter out trailing ALLREDUCE
        kernels = layer.kernels
        if kernels and kernels[-1].simplified_name == "ALLREDUCE":
            kernels = kernels[:-1]

        total_time = sum(k.duration_us for k in kernels)

        row = 1

        # Layer header
        ws.cell(row=row, column=1, value=f"Layer {layer.layer_idx} ({layer.layer_type.value}, {layer.stage.value})").font = Font(bold=True, size=12)
        row += 1
        ws.cell(row=row, column=1, value=f"Total Time: {total_time:.3f} us ({total_time/1000:.3f} ms)")
        row += 2

        # Kernel details header
        headers = ["#", "Duration (us)", "%", "Type", "Short Name", "Kernel Name"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
        row += 1

        # Kernel details
        for i, kernel in enumerate(kernels, 1):
            pct = f"{100.0 * kernel.duration_us / total_time:.1f}%" if total_time > 0 else "0.0%"
            ws.cell(row=row, column=1, value=i)
            ws.cell(row=row, column=2, value=round(kernel.duration_us, 3))
            ws.cell(row=row, column=3, value=pct)
            ws.cell(row=row, column=4, value=kernel.kernel_type.value)
            ws.cell(row=row, column=5, value=kernel.simplified_name)
            ws.cell(row=row, column=6, value=kernel.name)
            row += 1

        row += 1

        # Breakdown section
        ws.cell(row=row, column=1, value="=== BREAKDOWN ===").font = Font(bold=True)
        row += 1

        headers = ["Type", "Time (us)", "Percentage"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
        row += 1

        breakdown = {}
        for k in kernels:
            if k.kernel_type not in breakdown:
                breakdown[k.kernel_type] = 0.0
            breakdown[k.kernel_type] += k.duration_us

        for kt, time_us in sorted(breakdown.items(), key=lambda x: x[1], reverse=True):
            pct = f"{100.0 * time_us / total_time:.1f}%" if total_time > 0 else "0.0%"
            ws.cell(row=row, column=1, value=kt.value)
            ws.cell(row=row, column=2, value=round(time_us, 3))
            ws.cell(row=row, column=3, value=pct)
            row += 1

        # Auto-adjust column widths
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            # Limit kernel name column width
            if column == 'F':
                ws.column_dimensions[column].width = 80
            else:
                ws.column_dimensions[column].width = min(max_length + 2, 30)

    def _export_summary_csv(self, path: str) -> None:
        """Export summary statistics to CSV."""
        with open(path, "w", newline="") as f:
            writer = csv.writer(f)

            # Overall summary
            writer.writerow(["=== OVERALL SUMMARY ==="])
            writer.writerow(["metric", "value", "percentage"])
            total_time_us = self.result.total_time_us
            writer.writerow(["total_kernel_time_us", f"{total_time_us:.3f}", "100.0%"])

            for stage in [Stage.PREFILL, Stage.DECODE, Stage.UNKNOWN]:
                stats = self.result.per_stage_stats.get(stage, KernelStats())
                pct = f"{100.0 * stats.total_time_us / total_time_us:.1f}%" if total_time_us > 0 else "0.0%"
                writer.writerow([f"{stage.value}_time_us", f"{stats.total_time_us:.3f}", pct])

            writer.writerow([])

            # Type breakdown
            writer.writerow(["=== BREAKDOWN BY TYPE ==="])
            writer.writerow(["type", "count", "total_time_us", "avg_time_us", "percentage"])

            sorted_types = sorted(
                self.result.per_type_stats.items(),
                key=lambda x: x[1].total_time_us,
                reverse=True,
            )
            for kernel_type, stats in sorted_types:
                pct = f"{100.0 * stats.total_time_us / total_time_us:.1f}%" if total_time_us > 0 else "0.0%"
                writer.writerow([
                    kernel_type.value,
                    stats.count,
                    f"{stats.total_time_us:.3f}",
                    f"{stats.avg_time_us:.3f}",
                    pct,
                ])

            writer.writerow([])

            # Layer summary
            if self.result.layers:
                writer.writerow(["=== LAYER SUMMARY ==="])
                writer.writerow([
                    "layer_idx", "layer_type", "stage", "total_time_us",
                    "attention_time_us", "moe_time_us", "linear_time_us",
                    "communication_time_us", "quantization_time_us", "kernel_count"
                ])

                for layer in self.result.layers:
                    writer.writerow([
                        layer.layer_idx,
                        layer.layer_type.value,
                        layer.stage.value,
                        f"{layer.total_time_us:.3f}",
                        f"{layer.attention_time_us:.3f}",
                        f"{layer.moe_time_us:.3f}",
                        f"{layer.linear_time_us:.3f}",
                        f"{layer.communication_time_us:.3f}",
                        f"{layer.quantization_time_us:.3f}",
                        len(layer.kernels),
                    ])

        logger.info(f"Summary exported to: {path}")

    def _export_layer_detail_csv(self, layer: LayerEvent, path: str) -> None:
        """Export detailed kernel list for a single layer to CSV."""
        with open(path, "w", newline="") as f:
            writer = csv.writer(f)

            # Layer header
            writer.writerow([f"=== LAYER {layer.layer_idx} ({layer.layer_type.value}, {layer.stage.value}) ==="])
            writer.writerow([])

            # Filter out trailing ALLREDUCE
            kernels = layer.kernels
            if kernels and kernels[-1].simplified_name == "ALLREDUCE":
                kernels = kernels[:-1]

            total_time = sum(k.duration_us for k in kernels)

            # Kernel details
            writer.writerow(["index", "duration_us", "percentage", "type", "short_name", "kernel_name"])

            for i, kernel in enumerate(kernels, 1):
                pct = f"{100.0 * kernel.duration_us / total_time:.1f}%" if total_time > 0 else "0.0%"
                writer.writerow([
                    i,
                    f"{kernel.duration_us:.3f}",
                    pct,
                    kernel.kernel_type.value,
                    kernel.simplified_name,
                    kernel.name,
                ])

            writer.writerow([])

            # Layer total and breakdown
            writer.writerow(["=== LAYER TOTAL ==="])
            writer.writerow(["total_time_us", f"{total_time:.3f}"])
            writer.writerow([])

            writer.writerow(["=== BREAKDOWN ==="])
            writer.writerow(["type", "time_us", "percentage"])

            breakdown = {}
            for k in kernels:
                if k.kernel_type not in breakdown:
                    breakdown[k.kernel_type] = 0.0
                breakdown[k.kernel_type] += k.duration_us

            for kt, time_us in sorted(breakdown.items(), key=lambda x: x[1], reverse=True):
                pct = f"{100.0 * time_us / total_time:.1f}%" if total_time > 0 else "0.0%"
                writer.writerow([kt.value, f"{time_us:.3f}", pct])


def parse_layer_indices(layer_spec: str) -> Optional[List[int]]:
    """Parse layer specification string into list of indices.

    Supports formats:
    - "4,10,20" -> [4, 10, 20]
    - "4-10" -> [4, 5, 6, 7, 8, 9, 10]
    - "4-10,20,30-32" -> [4, 5, ..., 10, 20, 30, 31, 32]
    """
    if layer_spec is None or not layer_spec.strip():
        return []

    indices = []
    parts = layer_spec.split(",")
    for part in parts:
        part = part.strip()
        if "-" in part:
            start, end = part.split("-", 1)
            indices.extend(range(int(start), int(end) + 1))
        else:
            indices.append(int(part))

    return sorted(set(indices))


def main():
    parser = argparse.ArgumentParser(
        description="Analyze torch profiler trace files for kernel execution times.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  # Basic usage
  python -m sglang.srt.utils.trace_analyzer /path/to/trace.json.gz

  # Show top 30 kernels, decode stage only
  python -m sglang.srt.utils.trace_analyzer trace.json.gz --top-n 30 --stage decode

  # Layer-level analysis (terminal output)
  python -m sglang.srt.utils.trace_analyzer trace.json.gz --layer-analysis --show-layer-terminal

  # Show specific layer detail on terminal
  python -m sglang.srt.utils.trace_analyzer trace.json.gz --layer-analysis --layer-report 5 --show-layer-terminal

  # Export full analysis to Excel file (summary + layer details as sheets)
  python -m sglang.srt.utils.trace_analyzer trace.json.gz --export-csv analysis.xlsx --export-layers 4,10,20

  # Export layers 0-10 for prefill stage
  python -m sglang.srt.utils.trace_analyzer trace.json.gz --export-csv analysis.xlsx --export-layers 0-10 --stage prefill

  # Export to Excel without terminal output
  python -m sglang.srt.utils.trace_analyzer trace.json.gz --export-csv analysis.xlsx --export-layers 4,10

  # Debug specific layers
  python -m sglang.srt.utils.trace_analyzer trace.json.gz --layer-analysis --debug-layers 4,10
""",
    )

    parser.add_argument("trace_file", help="Path to trace file (.json.gz or .json)")
    parser.add_argument(
        "--top-n",
        type=int,
        default=20,
        help="Number of top kernels to show (default: 20)",
    )
    parser.add_argument(
        "--stage",
        choices=["prefill", "decode", "all"],
        default="all",
        help="Filter by stage (default: all)",
    )
    parser.add_argument(
        "--layer-analysis", action="store_true", help="Enable layer-level analysis"
    )
    parser.add_argument(
        "--layer-report", type=int, metavar="N", help="Show detailed report for layer N on terminal"
    )
    parser.add_argument(
        "--show-layer-terminal",
        action="store_true",
        help="Show layer analysis on terminal (default: False)",
    )
    parser.add_argument(
        "--export-csv",
        metavar="FILENAME",
        help="Export analysis to Excel file (.xlsx) with multiple sheets (Summary + Layer_N sheets)",
    )
    parser.add_argument(
        "--export-layers",
        metavar="LAYERS",
        help="Layers to include in CSV export: '4,10,20' or '4-10' (required with --export-csv)",
        default=None,
    )
    parser.add_argument(
        "--debug-layers",
        metavar="LAYERS",
        help="Print debug info for specific layers: '4,10,20' or '4-10'",
        default=None,
    )
    parser.add_argument(
        "--debug-top-n",
        type=int,
        default=20,
        help="Number of kernels to show in layer debug (default: 20)",
    )
    parser.add_argument(
        "--csv-stats", metavar="FILE", help="Export kernel stats to CSV file"
    )
    parser.add_argument(
        "--csv-events", metavar="FILE", help="Export all events to CSV file"
    )
    parser.add_argument(
        "--csv-layers", metavar="FILE", help="Export layer stats to CSV file"
    )
    parser.add_argument(
        "--grid-threshold",
        type=int,
        default=10000,
        help="Grid[0] threshold for decode heuristic (default: 10000)",
    )
    parser.add_argument(
        "--mtp-qseqlen-decode",
        action="store_true",
        help="Treat attention qseqlen>=2 kernels as decode for layer stage inference",
    )
    parser.add_argument(
        "--num-requests",
        type=int,
        default=1,
        metavar="N",
        help="Number of requests in trace for per-request TTFT calculation (default: 1)",
    )
    parser.add_argument(
        "--accept-length",
        type=float,
        default=1.0,
        metavar="L",
        help="MTP acceptance length for ITL correction (default: 1.0, no correction)",
    )
    parser.add_argument(
        "-v", "--verbose", action="store_true", help="Enable verbose logging"
    )

    args = parser.parse_args()

    # Configure logging
    log_level = logging.DEBUG if args.verbose else logging.INFO
    logging.basicConfig(
        level=log_level,
        format="%(asctime)s - %(levelname)s - %(message)s",
    )

    # Run analysis
    analyzer = TraceAnalyzer(
        grid_threshold=args.grid_threshold,
        mtp_qseqlen_decode=args.mtp_qseqlen_decode,
    )

    try:
        trace_data = analyzer.load_trace(args.trace_file)
    except FileNotFoundError:
        logger.error(f"Trace file not found: {args.trace_file}")
        sys.exit(1)
    except json.JSONDecodeError as e:
        logger.error(f"Invalid JSON in trace file: {e}")
        sys.exit(1)

    # Enable layer analysis if any layer-related option is requested
    detect_layers = (
        args.layer_analysis
        or args.layer_report is not None
        or args.csv_layers
        or args.export_csv
        or args.debug_layers
    )

    result = analyzer.analyze(trace_data, detect_layers=detect_layers)

    # Apply correction parameters for TTFT/ITL
    result.num_requests = args.num_requests
    result.accept_length = args.accept_length

    # Determine stage filter
    stage_enum = None
    if args.stage == "prefill":
        stage_enum = Stage.PREFILL
    elif args.stage == "decode":
        stage_enum = Stage.DECODE

    # Print report
    formatter = ReportFormatter(result)
    formatter.print_full_report(top_n=args.top_n, stage_filter=args.stage)

    # Print layer analysis on terminal only if --show-layer-terminal is set
    if detect_layers and args.show_layer_terminal:
        formatter.print_layer_summary(stage=stage_enum)

        if args.layer_report is not None:
            formatter.print_layer_detail(args.layer_report, stage=stage_enum)

    if detect_layers and args.debug_layers:
        debug_layers = parse_layer_indices(args.debug_layers)
        if not debug_layers:
            logger.error("--debug-layers must include at least one layer number.")
            sys.exit(1)
        for layer_idx in debug_layers:
            formatter.print_layer_debug(layer_idx, top_n=args.debug_top_n)

    # Export CSVs if requested
    exporter = CSVExporter(result)

    # Export full analysis to Excel file
    if args.export_csv:
        layer_indices = parse_layer_indices(args.export_layers)
        if not layer_indices:
            logger.error("--export-layers is required and must include at least one layer number.")
            sys.exit(1)

        # Save Excel file in the same folder as the input trace file
        trace_dir = os.path.dirname(os.path.abspath(args.trace_file))
        output_filename = args.export_csv if args.export_csv.endswith('.xlsx') else f"{args.export_csv}.xlsx"
        output_path = os.path.join(trace_dir, output_filename)

        try:
            created_file = exporter.export_full_analysis(
                output_path=output_path,
                layer_indices=layer_indices,
                stage_filter=stage_enum,
            )
            print(f"\nExported Excel file: {created_file}")
        except ImportError as e:
            logger.error(str(e))
            sys.exit(1)

    # Legacy CSV exports
    if args.csv_stats:
        exporter.export_kernel_stats(args.csv_stats)
    if args.csv_events:
        exporter.export_events(args.csv_events)
    if args.csv_layers:
        exporter.export_layers(args.csv_layers)


if __name__ == "__main__":
    main()
