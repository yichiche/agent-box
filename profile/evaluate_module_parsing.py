#!/usr/bin/env python3
"""
Trace Module Parsing Quality Evaluator

Reads Excel output from trace_module_analyzer.py and produces a quality score
(0-100) based on 4 structural rules (S1-S4) validating trace macro-structure.

This replaces evaluate_parsing.py for the new module-correlation-based analyzer.

Usage:
    python evaluate_module_parsing.py analysis.xlsx
    python evaluate_module_parsing.py analysis.xlsx --json
    python evaluate_module_parsing.py analysis.xlsx --threshold 85
"""

import argparse
import csv
import json
import math
import os
import re
import sys
from dataclasses import dataclass, field
from typing import Any, Dict, List, Optional, Tuple

import openpyxl


# ── Weights (edit here to tune scoring) ──────────────────────────────────────
W_S1_PHASE_COVERAGE       = 25   # S1: Phase coverage (prefill/decode detected)
W_S2_ARCHITECTURE_SIG     = 25   # S2: Architecture signature regularity
W_S3_INSTANCE_CONSISTENCY = 25   # S3: Instance count consistency across sibling types
W_S4_TIME_DISTRIBUTION    = 25   # S4: Time distribution consistency (low CV)


# ── Data model ───────────────────────────────────────────────────────────────

@dataclass
class ModuleTypeRecord:
    """From Overview sheet — one row per module type in hierarchy."""
    module_type: str
    depth: int
    count: int
    mean_us: float
    std_us: float
    min_us: float
    max_us: float
    total_us: float
    pct_of_parent: float
    top_kernel: str


@dataclass
class OverallStats:
    """From Summary sheet section 1."""
    total_kernel_time_us: float
    root_type_times: Dict[str, float]  # root module type -> total time


@dataclass
class CategoryBreakdown:
    """From Summary sheet section 2."""
    category: str
    count: int
    total_us: float
    avg_us: float
    percentage: float


@dataclass
class ModuleTreeEntry:
    """From Module Tree sheet."""
    module: str
    total_time_us: float
    kernels_ops: int
    breakdown: str
    phase: str


@dataclass
class DetailSheetInfo:
    """From per-module-type detail sheets."""
    module_type: str
    phase: str          # "pre" or "dec" (from sheet name suffix)
    kernel_count: int
    kernel_sum_us: float
    wall_time_us: float
    categories: Dict[str, float]  # category -> total duration
    kernels: List[Tuple[str, str, str, float, str]]  # (module_instance, input_dims, kernel_name, duration, category)


# ── Known sheets (not detail tabs) ──────────────────────────────────────────
_KNOWN_SHEETS = {"Summary", "Overview", "Module Tree", "GPU Kernels", "Model Info (WIP)"}


# ── Excel parsing functions ─────────────────────────────────────────────────

def _parse_summary_sheet(ws) -> Tuple[OverallStats, List[CategoryBreakdown]]:
    """Parse the Summary sheet into OverallStats and category breakdown."""
    rows = list(ws.iter_rows(values_only=True))

    total_kernel_time = 0.0
    root_type_times: Dict[str, float] = {}
    categories: List[CategoryBreakdown] = []

    # Section 1: Find "Total Kernel Time" row
    section = "header"
    cat_header_row = None

    for i, row in enumerate(rows):
        cell0 = str(row[0]).strip() if row[0] is not None else ""
        cell1 = row[1] if len(row) > 1 else None

        if section == "header":
            if cell0 == "Total Kernel Time":
                try:
                    total_kernel_time = float(cell1) if cell1 is not None else 0.0
                except (ValueError, TypeError):
                    total_kernel_time = 0.0
                section = "root_types"
                continue
            # Skip header row
            if cell0 in ("Metric", ""):
                continue

        elif section == "root_types":
            # Blank row or "Category" header signals end of root types
            if cell0 == "" or cell0 == "Category":
                if cell0 == "Category":
                    cat_header_row = i
                section = "categories"
                continue

            # Skip indented wrapper chain rows (they start with whitespace)
            if cell0.startswith(" ") or cell0.startswith("├") or cell0.startswith("└"):
                continue

            # Root module type row
            try:
                time_val = float(cell1) if cell1 is not None else 0.0
            except (ValueError, TypeError):
                time_val = 0.0
            root_type_times[cell0] = time_val

        elif section == "categories":
            if cat_header_row is None and cell0 == "Category":
                cat_header_row = i
                continue
            if cat_header_row is not None and cell0 and cell0 != "Category":
                try:
                    count = int(row[1]) if len(row) > 1 and row[1] is not None else 0
                    total_us = float(row[2]) if len(row) > 2 and row[2] is not None else 0.0
                    avg_us = float(row[3]) if len(row) > 3 and row[3] is not None else 0.0
                    pct_str = str(row[4]).replace("%", "") if len(row) > 4 and row[4] is not None else "0"
                    percentage = float(pct_str)
                except (ValueError, TypeError):
                    continue
                categories.append(CategoryBreakdown(
                    category=cell0, count=count, total_us=total_us,
                    avg_us=avg_us, percentage=percentage,
                ))

    overall = OverallStats(
        total_kernel_time_us=total_kernel_time,
        root_type_times=root_type_times,
    )
    return overall, categories


def _parse_overview_sheet(ws) -> List[ModuleTypeRecord]:
    """Parse the Overview sheet into a list of ModuleTypeRecord."""
    rows = list(ws.iter_rows(values_only=True))
    records = []

    if not rows:
        return records

    # Skip header row
    for row in rows[1:]:
        if row[0] is None:
            continue
        try:
            module_type = str(row[0]).strip()
            # Strip indent markers from module type name
            module_type = re.sub(r'^[\s├└─│]+', '', module_type).strip()
            if not module_type:
                continue

            depth = int(row[1]) if row[1] is not None else 0
            count = int(row[2]) if row[2] is not None else 0
            mean_us = float(row[3]) if row[3] is not None else 0.0
            std_us = float(row[4]) if row[4] is not None else 0.0
            min_us = float(row[5]) if row[5] is not None else 0.0
            max_us = float(row[6]) if row[6] is not None else 0.0
            total_us = float(row[7]) if row[7] is not None else 0.0
            pct_of_parent = float(row[8]) if row[8] is not None else 0.0
            top_kernel = str(row[9]) if len(row) > 9 and row[9] is not None else ""
        except (ValueError, TypeError, IndexError):
            continue

        records.append(ModuleTypeRecord(
            module_type=module_type, depth=depth, count=count,
            mean_us=mean_us, std_us=std_us, min_us=min_us, max_us=max_us,
            total_us=total_us, pct_of_parent=pct_of_parent, top_kernel=top_kernel,
        ))

    return records


def _parse_module_tree_sheet(ws) -> List[ModuleTreeEntry]:
    """Parse the Module Tree sheet."""
    rows = list(ws.iter_rows(values_only=True))
    entries = []

    if not rows:
        return entries

    for row in rows[1:]:
        if row[0] is None:
            continue
        try:
            module = str(row[0]).strip()
            total_time = float(row[1]) if len(row) > 1 and row[1] is not None else 0.0
            kernels_ops = int(row[2]) if len(row) > 2 and row[2] is not None else 0
            breakdown = str(row[3]) if len(row) > 3 and row[3] is not None else ""
            phase = str(row[4]) if len(row) > 4 and row[4] is not None else ""
        except (ValueError, TypeError):
            continue
        entries.append(ModuleTreeEntry(
            module=module, total_time_us=total_time, kernels_ops=kernels_ops,
            breakdown=breakdown, phase=phase,
        ))

    return entries


def _parse_detail_sheets(wb) -> List[DetailSheetInfo]:
    """Parse per-module-type detail sheets."""
    results = []

    for sheet_name in wb.sheetnames:
        if sheet_name in _KNOWN_SHEETS:
            continue

        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue

        # Parse sheet name for module_type and phase
        phase_match = re.search(r'\((\w+)\)\s*$', sheet_name)
        if phase_match:
            phase = phase_match.group(1)
            module_type = sheet_name[:phase_match.start()].strip()
        else:
            phase = ""
            module_type = sheet_name.strip()

        # Row 1: title line — "Name [phase] — N kernels"
        title_str = str(rows[0][0]) if rows[0][0] else ""
        kernel_count_match = re.search(r'(\d+)\s+kernels?', title_str)
        kernel_count = int(kernel_count_match.group(1)) if kernel_count_match else 0

        # Row 2: "Kernel sum: X us | Wall time: Y us | Overlap: Z us"
        kernel_sum_us = 0.0
        wall_time_us = 0.0
        if len(rows) > 1 and rows[1][0]:
            info_str = str(rows[1][0])
            sum_match = re.search(r'Kernel sum:\s*([\d,.]+)\s*us', info_str)
            wall_match = re.search(r'Wall time:\s*([\d,.]+)\s*us', info_str)
            if sum_match:
                kernel_sum_us = float(sum_match.group(1).replace(",", ""))
            if wall_match:
                wall_time_us = float(wall_match.group(1).replace(",", ""))

        # Parse category summary and kernel detail sections
        categories: Dict[str, float] = {}
        kernels: List[Tuple[str, str, str, float, str]] = []

        in_categories = False
        in_kernels = False

        for i, row in enumerate(rows[2:], start=2):
            cell0 = str(row[0]).strip() if row[0] is not None else ""

            # Detect category header
            if cell0 == "Category" and len(row) > 2:
                in_categories = True
                in_kernels = False
                continue

            # Detect kernel detail header
            if cell0 == "Module" and len(row) > 3:
                cell3 = str(row[2]).strip() if len(row) > 2 and row[2] is not None else ""
                if cell3 == "Kernel Name":
                    in_categories = False
                    in_kernels = True
                    continue

            if in_categories:
                if not cell0 or cell0 == "Total":
                    if cell0 == "Total":
                        in_categories = False
                    continue
                try:
                    dur = float(row[2]) if len(row) > 2 and row[2] is not None else 0.0
                    categories[cell0] = dur
                except (ValueError, TypeError):
                    pass

            elif in_kernels:
                if not cell0 or cell0.startswith("... truncated"):
                    continue
                try:
                    module_inst = cell0
                    input_dims = str(row[1]) if len(row) > 1 and row[1] is not None else ""
                    kernel_name = str(row[2]) if len(row) > 2 and row[2] is not None else ""
                    duration = float(row[3]) if len(row) > 3 and row[3] is not None else 0.0
                    category = str(row[5]) if len(row) > 5 and row[5] is not None else ""
                    kernels.append((module_inst, input_dims, kernel_name, duration, category))
                except (ValueError, TypeError):
                    continue

        results.append(DetailSheetInfo(
            module_type=module_type, phase=phase, kernel_count=kernel_count,
            kernel_sum_us=kernel_sum_us, wall_time_us=wall_time_us,
            categories=categories, kernels=kernels,
        ))

    return results


# ── Scoring rules ────────────────────────────────────────────────────────────

def _grade(score: float) -> str:
    if score >= 90:
        return "A"
    if score >= 80:
        return "B"
    if score >= 70:
        return "C"
    if score >= 60:
        return "D"
    return "F"


def _score_s1_phase_coverage(
    tree_entries: List[ModuleTreeEntry],
    overview_records: List[ModuleTypeRecord],
) -> Tuple[float, str]:
    """S1: Phase Coverage — check that phases are detected and root types exist."""
    phases_detected = set()
    for entry in tree_entries:
        phase = entry.phase.strip().lower()
        if phase:
            phases_detected.add(phase)

    # Check root types exist in overview
    root_types = [r for r in overview_records if r.depth == 0]
    n_root = len(root_types)

    if not phases_detected and n_root == 0:
        return 0.0, "no_phases_or_roots"

    score = 0.0
    # Having root types is fundamental
    if n_root > 0:
        score += 50.0

    # Having phases detected
    if "prefill" in phases_detected or "pre" in phases_detected:
        score += 25.0
    if "decode" in phases_detected or "dec" in phases_detected:
        score += 25.0

    # If no phases but root types exist, that's still partly OK (e.g. diffusion models)
    if not phases_detected and n_root > 0:
        score = 70.0  # Partial credit

    detail = f"phases_detected={{{','.join(sorted(phases_detected))}}}, root_types={n_root}"
    return min(score, 100.0), detail


def _score_s2_architecture_sig(
    overview_records: List[ModuleTypeRecord],
) -> Tuple[float, str]:
    """S2: Architecture Signature — regularity of module type hierarchy."""
    root_types = [r for r in overview_records if r.depth == 0]
    depth1_types = [r for r in overview_records if r.depth == 1]

    n_root = len(root_types)
    max_depth = max((r.depth for r in overview_records), default=0)

    if n_root == 0:
        return 0.0, "no_root_types"

    # Build pattern: root types with their first-level children
    pattern_parts = []
    for rt in root_types:
        pattern_parts.append(f"{rt.module_type}({rt.count}x)")

    # Score based on hierarchy regularity
    if n_root <= 3:
        score = 100.0
    elif n_root <= 5:
        score = 90.0
    elif n_root <= 10:
        score = 80.0
    else:
        score = 60.0

    # Bonus for having children (structured hierarchy)
    if depth1_types:
        has_children_root = sum(1 for d1 in depth1_types if d1.count > 0)
        if has_children_root > 0:
            score = min(score + 5, 100.0)

    detail = f"pattern=[{', '.join(pattern_parts[:5])}], root_types={n_root}, max_depth={max_depth}"
    return score, detail


def _score_s3_instance_consistency(
    overview_records: List[ModuleTypeRecord],
) -> Tuple[float, str]:
    """S3: Instance Count Consistency — sibling types should have consistent counts."""
    # Group depth-1 types (children of root types)
    depth1 = [r for r in overview_records if r.depth == 1 and r.count > 0]

    if len(depth1) < 2:
        return 100.0, "types_checked=0, consistent=0 (not enough types)"

    # Check if sibling counts are exact multiples of each other
    counts = [r.count for r in depth1]
    types_checked = len(depth1)
    consistent = 0
    inconsistent_types = []

    # Check each pair: counts should be multiples
    base_count = max(counts)
    for r in depth1:
        if base_count % r.count == 0 or r.count % base_count == 0:
            consistent += 1
        else:
            # Check if it's a reasonable multiple (within groups)
            is_multiple = False
            for c in counts:
                if c == r.count:
                    continue
                if c > 0 and (r.count % c == 0 or c % r.count == 0):
                    is_multiple = True
                    break
            if is_multiple:
                consistent += 1
            else:
                inconsistent_types.append(f"{r.module_type}({r.count})")

    if types_checked == 0:
        return 100.0, "types_checked=0"

    ratio = consistent / types_checked
    score = ratio * 100.0

    detail = (f"types_checked={types_checked}, consistent={consistent}"
              + (f", inconsistent_types=[{', '.join(inconsistent_types)}]"
                 if inconsistent_types else ""))
    return score, detail


def _score_s4_time_distribution(
    overview_records: List[ModuleTypeRecord],
) -> Tuple[float, str]:
    """S4: Time Distribution Consistency — low CV for repetitive module types."""
    # Filter to types with count >= 3 and depth >= 1
    eligible = [r for r in overview_records if r.count >= 3 and r.depth >= 1]

    if not eligible:
        # If no eligible types, check root types
        eligible = [r for r in overview_records if r.count >= 3 and r.depth == 0]

    if not eligible:
        return 100.0, "types_scored=0 (no repetitive types)"

    type_scores = []
    worst_cv = 0.0
    worst_type = ""
    cv_sum = 0.0

    for r in eligible:
        if r.mean_us > 0:
            cv = r.std_us / r.mean_us
        else:
            cv = 0.0

        if cv < 0.05:
            ts = 100.0
        elif cv < 0.10:
            ts = 80.0
        elif cv < 0.20:
            ts = 60.0
        elif cv < 0.50:
            ts = 40.0
        else:
            ts = 20.0

        # Weight by total time (more important modules count more)
        type_scores.append((ts, r.total_us))
        cv_sum += cv

        if cv > worst_cv:
            worst_cv = cv
            worst_type = r.module_type

    # Weighted average
    total_weight = sum(w for _, w in type_scores)
    if total_weight > 0:
        score = sum(s * w for s, w in type_scores) / total_weight
    else:
        score = sum(s for s, _ in type_scores) / len(type_scores)

    avg_cv = cv_sum / len(eligible) if eligible else 0.0

    detail = (f"types_scored={len(eligible)}, "
              f"worst_cv={worst_type}:{worst_cv:.3f}, "
              f"avg_cv={avg_cv:.3f}")
    return score, detail


# ── Per-group diagnostics ────────────────────────────────────────────────────

@dataclass
class GroupDiagnostic:
    """Diagnostic for one module type group."""
    module_type: str
    count: int
    depth: int
    time_consistency_score: float
    time_consistency_detail: str
    kernel_pattern_score: float
    kernel_pattern_detail: str
    composite_score: float


def _compute_group_diagnostics(
    overview_records: List[ModuleTypeRecord],
    detail_sheets: List[DetailSheetInfo],
) -> List[GroupDiagnostic]:
    """Compute per-group diagnostics for eligible module types."""
    diagnostics = []

    # Eligible: depth >= 1, count >= 3
    eligible = [r for r in overview_records if r.count >= 3 and r.depth >= 1]

    # Build detail sheet lookup
    detail_by_type: Dict[str, List[DetailSheetInfo]] = {}
    for ds in detail_sheets:
        detail_by_type.setdefault(ds.module_type, []).append(ds)

    for r in eligible:
        # Time consistency from Overview
        if r.mean_us > 0:
            cv = r.std_us / r.mean_us
        else:
            cv = 0.0

        if cv < 0.05:
            time_score = 100.0
        elif cv < 0.10:
            time_score = 80.0
        elif cv < 0.20:
            time_score = 60.0
        else:
            time_score = max(0.0, 100.0 - cv * 200)

        time_detail = f"cv={cv:.3f}, mean={r.mean_us:.1f}, std={r.std_us:.1f}"

        # Kernel pattern consistency from detail sheets
        sheets = detail_by_type.get(r.module_type, [])
        kernel_score = 100.0
        kernel_detail = "no_detail_sheet"

        if sheets:
            # Check kernel sequences across module instances within detail sheet
            for ds in sheets:
                # Group kernels by module instance
                by_instance: Dict[str, List[str]] = {}
                for mod_inst, _, kname, _, _ in ds.kernels:
                    by_instance.setdefault(mod_inst, []).append(kname)

                if len(by_instance) >= 2:
                    seqs = list(by_instance.values())
                    ref_seq = seqs[0]
                    matches = sum(1 for s in seqs[1:] if s == ref_seq)
                    kernel_score = (1 + matches) / len(seqs) * 100
                    kernel_detail = f"instances={len(seqs)}, matching={1 + matches}"
                    break

        composite = time_score * 0.7 + kernel_score * 0.3

        diagnostics.append(GroupDiagnostic(
            module_type=r.module_type, count=r.count, depth=r.depth,
            time_consistency_score=time_score, time_consistency_detail=time_detail,
            kernel_pattern_score=kernel_score, kernel_pattern_detail=kernel_detail,
            composite_score=composite,
        ))

    return diagnostics


# ── Main evaluator ───────────────────────────────────────────────────────────

class ModuleParsingEvaluator:
    """Evaluates trace_module_analyzer.py Excel output quality."""

    def __init__(self, xlsx_path: str):
        self.xlsx_path = xlsx_path
        self.wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)

        # Parse all sheets
        self.overall_stats = None
        self.categories = []
        self.overview_records = []
        self.tree_entries = []
        self.detail_sheets = []

        if "Summary" in self.wb.sheetnames:
            self.overall_stats, self.categories = _parse_summary_sheet(self.wb["Summary"])
        if "Overview" in self.wb.sheetnames:
            self.overview_records = _parse_overview_sheet(self.wb["Overview"])
        if "Module Tree" in self.wb.sheetnames:
            self.tree_entries = _parse_module_tree_sheet(self.wb["Module Tree"])

        self.detail_sheets = _parse_detail_sheets(self.wb)
        self.wb.close()

        # Compute scores
        self.scores = self._compute_scores()
        self.diagnostics = _compute_group_diagnostics(
            self.overview_records, self.detail_sheets)

    def _compute_scores(self) -> Dict[str, Dict[str, Any]]:
        """Compute S1-S4 structural rule scores."""
        s1_score, s1_detail = _score_s1_phase_coverage(
            self.tree_entries, self.overview_records)
        s2_score, s2_detail = _score_s2_architecture_sig(self.overview_records)
        s3_score, s3_detail = _score_s3_instance_consistency(self.overview_records)
        s4_score, s4_detail = _score_s4_time_distribution(self.overview_records)

        return {
            "s1_phase_coverage": {
                "score": round(s1_score, 1),
                "grade": _grade(s1_score),
                "weight": W_S1_PHASE_COVERAGE,
                "detail": s1_detail,
            },
            "s2_architecture_sig": {
                "score": round(s2_score, 1),
                "grade": _grade(s2_score),
                "weight": W_S2_ARCHITECTURE_SIG,
                "detail": s2_detail,
            },
            "s3_instance_consistency": {
                "score": round(s3_score, 1),
                "grade": _grade(s3_score),
                "weight": W_S3_INSTANCE_CONSISTENCY,
                "detail": s3_detail,
            },
            "s4_time_distribution": {
                "score": round(s4_score, 1),
                "grade": _grade(s4_score),
                "weight": W_S4_TIME_DISTRIBUTION,
                "detail": s4_detail,
            },
        }

    @property
    def composite_score(self) -> float:
        """Weighted composite score (0-100)."""
        total = 0.0
        for rule in self.scores.values():
            total += rule["score"] * rule["weight"] / 100.0
        return round(total, 1)

    def export_csv(self, output_dir: str = ".") -> str:
        """Export evaluation summary as CSV. Returns the output path."""
        csv_path = os.path.join(output_dir, "evaluation_summary.csv")

        rows = []

        # Structural rules
        for key, data in self.scores.items():
            rows.append({
                "section": "structural_rules",
                "metric": key,
                "score": data["score"],
                "grade": data["grade"],
                "detail": data["detail"],
            })

        # Group diagnostics
        for diag in self.diagnostics:
            rows.append({
                "section": "group_diagnostics",
                "metric": diag.module_type,
                "score": round(diag.composite_score, 1),
                "grade": _grade(diag.composite_score),
                "detail": (f"count={diag.count}, depth={diag.depth}, "
                          f"time_cv={diag.time_consistency_detail}, "
                          f"kernel={diag.kernel_pattern_detail}"),
            })

        # Composite
        rows.append({
            "section": "composite",
            "metric": "overall_score",
            "score": self.composite_score,
            "grade": _grade(self.composite_score),
            "detail": f"weighted_sum_of_s1-s4",
        })

        with open(csv_path, "w", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(
                f, fieldnames=["section", "metric", "score", "grade", "detail"])
            writer.writeheader()
            writer.writerows(rows)

        return csv_path

    def to_json(self) -> Dict[str, Any]:
        """Return JSON-serializable evaluation result."""
        return {
            "file": self.xlsx_path,
            "composite_score": self.composite_score,
            "composite_grade": _grade(self.composite_score),
            "structural_rules": self.scores,
            "group_diagnostics": [
                {
                    "module_type": d.module_type,
                    "count": d.count,
                    "depth": d.depth,
                    "time_consistency": {
                        "score": round(d.time_consistency_score, 1),
                        "detail": d.time_consistency_detail,
                    },
                    "kernel_pattern": {
                        "score": round(d.kernel_pattern_score, 1),
                        "detail": d.kernel_pattern_detail,
                    },
                    "composite_score": round(d.composite_score, 1),
                }
                for d in self.diagnostics
            ],
        }

    def print_report(self) -> None:
        """Print a human-readable evaluation report."""
        sep = "=" * 60
        print()
        print(sep)
        print("  Trace Module Parsing Quality Report")
        print(f"  File: {self.xlsx_path}")
        print(sep)
        print()

        # Structural rules
        print("--- Structural Rules ---")
        print(f"  {'Rule':<28} | {'Score':>6} | {'Grade':>5} | Detail")
        print(f"  {'-'*28}-+-{'-'*6}-+-{'-'*5}-+-{'-'*40}")
        for key, data in self.scores.items():
            print(f"  {key:<28} | {data['score']:>6.1f} | {data['grade']:>5} | {data['detail']}")

        # Group diagnostics
        if self.diagnostics:
            print()
            print("--- Group Diagnostics ---")
            print(f"  {'Module Type':<30} | {'Count':>5} | {'Time':>6} | {'Kernel':>6} | {'Score':>6}")
            print(f"  {'-'*30}-+-{'-'*5}-+-{'-'*6}-+-{'-'*6}-+-{'-'*6}")
            for d in self.diagnostics:
                print(f"  {d.module_type:<30} | {d.count:>5} | "
                      f"{d.time_consistency_score:>6.1f} | "
                      f"{d.kernel_pattern_score:>6.1f} | "
                      f"{d.composite_score:>6.1f}")

        print()
        print(f"  Composite Score: {self.composite_score:.1f} ({_grade(self.composite_score)})")
        print()


# ── CLI ──────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description="Evaluate trace_module_analyzer.py Excel output quality.",
    )
    parser.add_argument("xlsx_path", help="Path to analysis Excel file")
    parser.add_argument("--json", action="store_true", help="Output JSON instead of report")
    parser.add_argument(
        "--threshold", type=float, default=None,
        help="Exit with code 1 if composite score < threshold",
    )
    args = parser.parse_args()

    if not os.path.isfile(args.xlsx_path):
        print(f"ERROR: File not found: {args.xlsx_path}", file=sys.stderr)
        sys.exit(1)

    evaluator = ModuleParsingEvaluator(args.xlsx_path)

    # Always export CSV to the directory of the input file
    csv_dir = os.path.dirname(os.path.abspath(args.xlsx_path))
    csv_path = evaluator.export_csv(csv_dir)

    if args.json:
        print(json.dumps(evaluator.to_json(), indent=2))
    else:
        evaluator.print_report()
        print(f"  CSV saved: {csv_path}")

    if args.threshold is not None and evaluator.composite_score < args.threshold:
        sys.exit(1)


if __name__ == "__main__":
    main()
