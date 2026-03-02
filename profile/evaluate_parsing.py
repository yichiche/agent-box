#!/usr/bin/env python3
"""
Trace Parsing Quality Evaluator

Reads Excel output from trace_analyzer.py and produces a quality score (0-100)
based on 4 structural rules (S1-S4) validating trace macro-structure.

Usage:
    python evaluate_parsing.py analysis.xlsx
    python evaluate_parsing.py analysis.xlsx --json
    python evaluate_parsing.py analysis.xlsx --threshold 85
"""

import argparse
import collections
import json
import math
import re
import sys
from dataclasses import dataclass, field
from typing import Any, Dict, List, Optional, Tuple

import openpyxl


# ── Weights (edit here to tune scoring) ──────────────────────────────────────
# 4 structural rules, total weight = 100 (absolute, no normalization).

W_S1_PREFILL_ORDERING     = 25   # S1: No decode before first prefill
W_S2_ARCHITECTURE_SIG     = 25   # S2: First prefill round matches model pattern
W_S3_ROUND_CONSISTENCY    = 25   # S3: Decode rounds all have same layer count
W_S4_TYPE_SEQUENCE        = 25   # S4: Type pattern repeats per round


# ── Data model ───────────────────────────────────────────────────────────────

@dataclass
class LayerSummaryRecord:
    """One row from the Summary sheet layer table."""
    layer_idx: int
    layer_type: str
    stage: str
    total_time_us: float
    kernel_count: int
    attention_time_us: float = 0.0
    moe_time_us: float = 0.0
    linear_time_us: float = 0.0
    comm_time_us: float = 0.0
    quant_time_us: float = 0.0


@dataclass
class LayerKernelPattern:
    """Kernel sequence from a Layer_N sheet."""
    layer_idx: int
    layer_type: str
    stage: str
    short_names: Tuple[str, ...]


@dataclass
class OverallStats:
    """Overall time stats parsed from Summary sheet."""
    total_time_us: float
    prefill_time_us: float
    decode_time_us: float
    unknown_time_us: float


@dataclass
class GroupKey:
    """(stage, layer_type) identifier."""
    stage: str
    layer_type: str

    def __hash__(self):
        return hash((self.stage, self.layer_type))

    def __eq__(self, other):
        return isinstance(other, GroupKey) and self.stage == other.stage and self.layer_type == other.layer_type

    def __lt__(self, other):
        return (self.stage, self.layer_type) < (other.stage, other.layer_type)

    def __repr__(self):
        return f"({self.stage}, {self.layer_type})"


@dataclass
class ModelArchitectureConfig:
    """Optional model hints for architecture-aware scoring."""
    num_hidden_layers: Optional[int] = None
    expected_prefill_pattern: Optional[List[Tuple[int, str]]] = None  # [(count, type), ...]
    expected_decode_pattern: Optional[List[Tuple[int, str]]] = None   # [(count, type), ...]


@dataclass
class StructuralMetrics:
    """Results for all 4 structural rules."""
    # S1: Prefill-Before-Decode Ordering
    s1_score: float = 0.0
    s1_details: Dict[str, Any] = field(default_factory=dict)

    # S2: Architecture Signature
    s2_score: float = 0.0
    s2_details: Dict[str, Any] = field(default_factory=dict)

    # S3: Consistent Round Size
    s3_score: float = 0.0
    s3_details: Dict[str, Any] = field(default_factory=dict)

    # S4: Layer Type Sequence Repeatability
    s4_score: float = 0.0
    s4_details: Dict[str, Any] = field(default_factory=dict)


@dataclass
class GroupMetrics:
    """All computed metrics for one group (diagnostic)."""
    key: GroupKey
    layer_count: int
    skipped: bool = False
    skip_reason: str = ""

    # Time consistency (diagnostic)
    time_cv: float = 0.0
    time_score: float = 0.0

    # Kernel count consistency (diagnostic)
    kernel_count_mode: int = 0
    kernel_count_mode_pct: float = 0.0
    kernel_count_distinct: int = 0
    kernel_count_score: float = 0.0

    # Kernel pattern consistency (diagnostic)
    pattern_mode_pct: float = 0.0
    pattern_available: bool = False
    pattern_score: float = 0.0

    # Outlier info (diagnostic)
    outlier_count: int = 0
    outlier_pct: float = 0.0
    outlier_score: float = 0.0
    outlier_indices: List[int] = field(default_factory=list)

    # Composite (diagnostic)
    composite_score: float = 0.0

    # Advisory notes
    notes: List[str] = field(default_factory=list)


@dataclass
class EvaluationReport:
    """Complete evaluation result."""
    filepath: str
    total_layers: int
    group_count: int

    # Structural rules
    structural_metrics: StructuralMetrics
    structural_score: float

    # Per-group (diagnostics)
    group_metrics: List[GroupMetrics]

    # Overall
    overall_score: float
    grade: str
    passed: bool
    threshold: float

    # All outlier indices (diagnostic)
    all_outlier_indices: List[int]


# ── Scoring helpers ──────────────────────────────────────────────────────────

def _score_stage_coverage(pct: float) -> float:
    """Stage classification coverage percentage -> score."""
    if pct >= 100.0 - 1e-9:
        return 100.0
    if pct >= 98.0:
        return 80.0
    if pct >= 90.0:
        return 60.0
    # Linear decay from 60 at 90% to 0 at 0%
    return max(0.0, 60.0 * pct / 90.0)


def _score_time_cv(cv: float) -> float:
    """Time CV scoring: trimmed CV -> score (diagnostic)."""
    if cv <= 0.05:
        return 100.0
    if cv <= 0.10:
        return 80.0
    if cv <= 0.20:
        return 60.0
    # Linear decay from 60 at 0.20 to 0 at 1.0
    return max(0.0, 60.0 * (1.0 - cv) / 0.80)


def _score_mode_match(pct: float) -> float:
    """Mode match percentage -> score."""
    if pct >= 95.0 - 1e-9:
        return 100.0
    if pct >= 80.0:
        return 80.0
    if pct >= 60.0:
        return 60.0
    # Linear decay from 60 at 60% to 0 at 0%
    return max(0.0, 60.0 * pct / 60.0)


def _score_outlier_pct(pct: float) -> float:
    """Outlier percentage -> score (diagnostic)."""
    if pct < 2.0:
        return 100.0
    if pct < 5.0:
        return 80.0
    if pct < 10.0:
        return 60.0
    # Linear decay from 60 at 10% to 0 at 100%
    return max(0.0, 60.0 * (100.0 - pct) / 90.0)


def _grade(score: float) -> str:
    if score >= 90:
        return "A"
    if score >= 75:
        return "B"
    if score >= 60:
        return "C"
    return "F"


# ── Statistics helpers ───────────────────────────────────────────────────────

def _trimmed_cv(values: List[float]) -> Tuple[float, List[int]]:
    """Compute trimmed coefficient of variation after IQR outlier removal.

    Returns (cv, outlier_indices_in_original_list).
    """
    if len(values) < 2:
        return 0.0, []

    indexed = sorted(enumerate(values), key=lambda x: x[1])
    sorted_vals = [v for _, v in indexed]
    n = len(sorted_vals)

    q1_idx = n // 4
    q3_idx = (3 * n) // 4
    q1 = sorted_vals[q1_idx]
    q3 = sorted_vals[q3_idx]
    iqr = q3 - q1

    lower = q1 - 1.5 * iqr
    upper = q3 + 1.5 * iqr

    inliers = []
    outlier_indices = []
    for orig_idx, val in indexed:
        if val < lower or val > upper:
            outlier_indices.append(orig_idx)
        else:
            inliers.append(val)

    if len(inliers) < 2:
        # All outliers — use original values
        mean = sum(values) / len(values)
        if mean == 0:
            return 0.0, outlier_indices
        std = math.sqrt(sum((v - mean) ** 2 for v in values) / len(values))
        return std / mean, outlier_indices

    mean = sum(inliers) / len(inliers)
    if mean == 0:
        return 0.0, outlier_indices
    std = math.sqrt(sum((v - mean) ** 2 for v in inliers) / len(inliers))
    return std / mean, outlier_indices


def _median(values: List[float]) -> float:
    """Compute the median of a list of values."""
    if not values:
        return 0.0
    s = sorted(values)
    n = len(s)
    if n % 2 == 1:
        return s[n // 2]
    return (s[n // 2 - 1] + s[n // 2]) / 2.0


def _segment_into_rounds(records: List[LayerSummaryRecord]) -> List[Tuple[str, List[LayerSummaryRecord]]]:
    """Segment a layer sequence into rounds by stage transitions.

    A round is a contiguous block of layers with the same stage.
    Returns [(stage, [records]), ...] in order.
    """
    if not records:
        return []

    rounds: List[Tuple[str, List[LayerSummaryRecord]]] = []
    current_stage = records[0].stage
    current_group: List[LayerSummaryRecord] = [records[0]]

    for rec in records[1:]:
        if rec.stage != current_stage:
            rounds.append((current_stage, current_group))
            current_stage = rec.stage
            current_group = [rec]
        else:
            current_group.append(rec)

    rounds.append((current_stage, current_group))
    return rounds


def _detect_round_size(rounds: List[Tuple[str, List[LayerSummaryRecord]]], stage: str) -> int:
    """Detect the mode round size for a given stage."""
    sizes = [len(recs) for s, recs in rounds if s == stage]
    if not sizes:
        return 0
    counter = collections.Counter(sizes)
    return counter.most_common(1)[0][0]


# ── Excel parsing ────────────────────────────────────────────────────────────

_HEADER_RE = re.compile(r"Layer\s+\d+\s+\(([^,]+),\s*(\w+)\)")


def _parse_summary_sheet(ws) -> Tuple[OverallStats, List[LayerSummaryRecord]]:
    """Parse the Summary sheet for overall stats and layer records."""
    rows = list(ws.iter_rows(values_only=True))

    total_time = 0.0
    prefill_time = 0.0
    decode_time = 0.0
    unknown_time = 0.0
    records = []

    # Parse overall summary section
    in_overall = False
    in_layers = False
    layer_header_row = -1

    for i, row in enumerate(rows):
        cell0 = str(row[0]) if row[0] is not None else ""

        if "OVERALL SUMMARY" in cell0:
            in_overall = True
            in_layers = False
            continue
        if "LAYER SUMMARY" in cell0:
            in_overall = False
            in_layers = True
            continue
        if "BREAKDOWN BY TYPE" in cell0:
            in_overall = False
            in_layers = False
            continue

        if in_overall:
            if cell0 == "Total Kernel Time":
                total_time = float(row[1]) if row[1] is not None else 0.0
            elif cell0 == "Prefill Time":
                prefill_time = float(row[1]) if row[1] is not None else 0.0
            elif cell0 == "Decode Time":
                decode_time = float(row[1]) if row[1] is not None else 0.0
            elif cell0 == "Unknown Time":
                unknown_time = float(row[1]) if row[1] is not None else 0.0

        if in_layers:
            if cell0 == "Layer":
                # This is the header row
                layer_header_row = i
                continue
            if layer_header_row >= 0 and row[0] is not None:
                try:
                    layer_idx = int(row[0])
                    layer_type = str(row[1]) if row[1] else "unknown"
                    stage = str(row[2]) if row[2] else "unknown"
                    total_us = float(row[3]) if row[3] is not None else 0.0
                    attn_us = float(row[4]) if row[4] is not None else 0.0
                    moe_us = float(row[5]) if row[5] is not None else 0.0
                    linear_us = float(row[6]) if row[6] is not None else 0.0
                    comm_us = float(row[7]) if row[7] is not None else 0.0
                    quant_us = float(row[8]) if row[8] is not None else 0.0
                    kernels = int(row[9]) if row[9] is not None else 0
                    records.append(LayerSummaryRecord(
                        layer_idx=layer_idx,
                        layer_type=layer_type,
                        stage=stage,
                        total_time_us=total_us,
                        kernel_count=kernels,
                        attention_time_us=attn_us,
                        moe_time_us=moe_us,
                        linear_time_us=linear_us,
                        comm_time_us=comm_us,
                        quant_time_us=quant_us,
                    ))
                except (ValueError, TypeError, IndexError):
                    pass

    overall = OverallStats(total_time, prefill_time, decode_time, unknown_time)
    return overall, records


def _parse_layer_sheets(wb) -> List[LayerKernelPattern]:
    """Parse Layer_N sheets for kernel patterns."""
    patterns = []

    for sheet_name in wb.sheetnames:
        if not sheet_name.startswith("Layer_"):
            continue

        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue

        header_text = str(rows[0][0]) if rows[0][0] else ""
        m = _HEADER_RE.match(header_text)
        if not m:
            continue

        layer_type = m.group(1).strip()
        stage = m.group(2).strip()
        layer_idx = int(sheet_name.split("_")[1])

        # Find kernel data table
        data_start = None
        for i, row in enumerate(rows):
            if row[0] == "#" and len(row) > 4 and row[4] == "Short Name":
                data_start = i + 1
                break

        if data_start is None:
            continue

        short_names = []
        for row in rows[data_start:]:
            if row[0] is None or str(row[0]).startswith("=== BREAKDOWN"):
                break
            sn = str(row[4]) if row[4] else ""
            short_names.append(sn)

        patterns.append(LayerKernelPattern(
            layer_idx=layer_idx,
            layer_type=layer_type,
            stage=stage,
            short_names=tuple(short_names),
        ))

    return patterns


# ── Main evaluator class ─────────────────────────────────────────────────────

class ParsingEvaluator:
    """Evaluates trace parsing quality from Excel output."""

    MIN_GROUP_SIZE = 3  # Skip groups with fewer layers

    def __init__(self, threshold: float = 70.0, config: Optional[ModelArchitectureConfig] = None):
        self.threshold = threshold
        self.config = config or ModelArchitectureConfig()

    # ── Data loading ─────────────────────────────────────────────────────

    def load_from_excel(self, filepath: str) -> Tuple[OverallStats, List[LayerSummaryRecord], List[LayerKernelPattern]]:
        """Load data from an Excel file."""
        wb = openpyxl.load_workbook(filepath, read_only=True)
        ws_summary = wb["Summary"]
        overall, records = _parse_summary_sheet(ws_summary)
        patterns = _parse_layer_sheets(wb)
        wb.close()
        return overall, records, patterns

    def load_from_analysis_result(self, result) -> Tuple[OverallStats, List[LayerSummaryRecord], List[LayerKernelPattern]]:
        """Load data from an AnalysisResult object (for test integration)."""
        from trace_analyzer import Stage

        total_time = result.total_time_us
        prefill_stats = result.per_stage_stats.get(Stage.PREFILL, None)
        decode_stats = result.per_stage_stats.get(Stage.DECODE, None)
        unknown_stats = result.per_stage_stats.get(Stage.UNKNOWN, None)

        overall = OverallStats(
            total_time_us=total_time,
            prefill_time_us=prefill_stats.total_time_us if prefill_stats else 0.0,
            decode_time_us=decode_stats.total_time_us if decode_stats else 0.0,
            unknown_time_us=unknown_stats.total_time_us if unknown_stats else 0.0,
        )

        records = []
        patterns = []
        for layer in result.layers:
            records.append(LayerSummaryRecord(
                layer_idx=layer.layer_idx,
                layer_type=layer.layer_type.value,
                stage=layer.stage.value,
                total_time_us=layer.total_time_us,
                kernel_count=len(layer.kernels),
                attention_time_us=layer.attention_time_us,
                moe_time_us=layer.moe_time_us,
                linear_time_us=layer.linear_time_us,
                comm_time_us=layer.communication_time_us,
                quant_time_us=layer.quantization_time_us,
            ))
            # Build kernel pattern from kernel events
            short_names = tuple(k.simplified_name for k in layer.kernels)
            patterns.append(LayerKernelPattern(
                layer_idx=layer.layer_idx,
                layer_type=layer.layer_type.value,
                stage=layer.stage.value,
                short_names=short_names,
            ))

        return overall, records, patterns

    # ── Grouping ─────────────────────────────────────────────────────────

    def group_layers(
        self,
        records: List[LayerSummaryRecord],
        patterns: List[LayerKernelPattern],
    ) -> Dict[GroupKey, Tuple[List[LayerSummaryRecord], List[LayerKernelPattern]]]:
        """Group records and patterns by (stage, layer_type)."""
        groups: Dict[GroupKey, Tuple[List[LayerSummaryRecord], List[LayerKernelPattern]]] = {}

        for rec in records:
            key = GroupKey(rec.stage, rec.layer_type)
            if key not in groups:
                groups[key] = ([], [])
            groups[key][0].append(rec)

        # Build pattern lookup by layer_idx for matching
        pattern_map = {p.layer_idx: p for p in patterns}
        for key, (recs, pats) in groups.items():
            for rec in recs:
                if rec.layer_idx in pattern_map:
                    p = pattern_map[rec.layer_idx]
                    if p.stage == key.stage and p.layer_type == key.layer_type:
                        pats.append(p)

        return groups

    # ── S1: Prefill-Before-Decode Ordering ──────────────────────────────

    def compute_prefill_ordering(self, records: List[LayerSummaryRecord]) -> Tuple[float, Dict[str, Any]]:
        """S1: Check that no decode layer appears before the first prefill layer."""
        if not records:
            return 100.0, {"first_prefill_idx": None, "first_decode_idx": None, "violations": 0}

        first_prefill_pos = None
        first_decode_pos = None
        for i, rec in enumerate(records):
            if rec.stage == "prefill" and first_prefill_pos is None:
                first_prefill_pos = i
            if rec.stage == "decode" and first_decode_pos is None:
                first_decode_pos = i

        violations = 0
        if first_decode_pos is not None and first_prefill_pos is not None:
            if first_decode_pos < first_prefill_pos:
                # Count how many decode layers appear before first prefill
                for i in range(first_prefill_pos):
                    if records[i].stage == "decode":
                        violations += 1

        if violations == 0:
            score = 100.0
        else:
            # Penalize proportionally: each violation costs up to 20 points
            score = max(0.0, 100.0 - violations * 20.0)

        details = {
            "first_prefill_idx": records[first_prefill_pos].layer_idx if first_prefill_pos is not None else None,
            "first_decode_idx": records[first_decode_pos].layer_idx if first_decode_pos is not None else None,
            "violations": violations,
        }
        return score, details

    # ── S2: Architecture Signature ──────────────────────────────────────

    def compute_architecture_signature(
        self,
        records: List[LayerSummaryRecord],
        config: Optional[ModelArchitectureConfig] = None,
    ) -> Tuple[float, Dict[str, Any]]:
        """S2: Extract first prefill round's type pattern and optionally match config."""
        cfg = config or self.config
        rounds = _segment_into_rounds(records)

        # Find first prefill round
        first_prefill_round = None
        for stage, recs in rounds:
            if stage == "prefill":
                first_prefill_round = recs
                break

        if first_prefill_round is None:
            return 0.0, {"detected_pattern": [], "matched": False, "reason": "no prefill round found"}

        # Extract type pattern as run-length encoding
        detected = _run_length_encode([r.layer_type for r in first_prefill_round])

        details: Dict[str, Any] = {
            "detected_pattern": detected,
            "round_size": len(first_prefill_round),
        }

        if cfg.expected_prefill_pattern is not None:
            matched = detected == cfg.expected_prefill_pattern
            details["matched"] = matched
            details["expected_pattern"] = cfg.expected_prefill_pattern
            score = 100.0 if matched else 30.0
        else:
            # Auto-detect: score based on whether pattern is regular
            # A good pattern has few distinct types and clear repetition
            n_distinct_types = len(set(t for _, t in detected))
            total_layers = sum(c for c, _ in detected)
            if total_layers >= 10 and n_distinct_types <= 10:
                score = 100.0
            elif total_layers >= 5:
                score = 80.0
            else:
                score = 60.0
            details["matched"] = True  # auto-detected is considered a match
            details["auto_detected"] = True

        return score, details

    # ── S3: Consistent Round Size ───────────────────────────────────────

    def compute_round_consistency(self, records: List[LayerSummaryRecord]) -> Tuple[float, Dict[str, Any]]:
        """S3: Check that all decode rounds have the same layer count."""
        rounds = _segment_into_rounds(records)
        decode_sizes = [len(recs) for stage, recs in rounds if stage == "decode"]

        if not decode_sizes:
            return 100.0, {"decode_rounds": 0, "mode_size": 0, "match_pct": 100.0}

        counter = collections.Counter(decode_sizes)
        mode_size, mode_freq = counter.most_common(1)[0]
        match_pct = 100.0 * mode_freq / len(decode_sizes)

        score = _score_mode_match(match_pct)

        details = {
            "decode_rounds": len(decode_sizes),
            "mode_size": mode_size,
            "match_pct": round(match_pct, 1),
            "distinct_sizes": len(counter),
            "size_distribution": dict(counter),
        }
        return score, details

    # ── S4: Layer Type Sequence Repeatability ───────────────────────────

    def compute_type_sequence_repeatability(self, records: List[LayerSummaryRecord]) -> Tuple[float, Dict[str, Any]]:
        """S4: Check that each round's type sequence matches the dominant pattern."""
        rounds = _segment_into_rounds(records)
        if not rounds:
            return 100.0, {"rounds": 0, "match_pct": 100.0}

        # Build type sequences per round (for all stages combined)
        round_patterns = []
        for stage, recs in rounds:
            pattern = tuple(r.layer_type for r in recs)
            round_patterns.append((stage, pattern))

        # Group by stage and check pattern consistency within each stage
        stage_patterns: Dict[str, List[Tuple[str, ...]]] = {}
        for stage, pattern in round_patterns:
            stage_patterns.setdefault(stage, []).append(pattern)

        total_rounds = 0
        matching_rounds = 0
        for stage, pats in stage_patterns.items():
            if len(pats) < 2:
                # Single round — trivially matches
                total_rounds += len(pats)
                matching_rounds += len(pats)
                continue
            counter = collections.Counter(pats)
            mode_pat, mode_freq = counter.most_common(1)[0]
            total_rounds += len(pats)
            matching_rounds += mode_freq

        match_pct = 100.0 * matching_rounds / total_rounds if total_rounds > 0 else 100.0
        score = _score_mode_match(match_pct)

        details = {
            "rounds": total_rounds,
            "matching_rounds": matching_rounds,
            "match_pct": round(match_pct, 1),
        }
        return score, details

    # ── S6: Stage Transition Regularity ─────────────────────────────────

    def compute_transition_regularity(self, records: List[LayerSummaryRecord]) -> Tuple[float, Dict[str, Any]]:
        """S6: Check that stage transitions follow a regular pattern.

        Expected: chunked prefill blocks followed by uniform decode blocks,
        with clean transitions (no interleaving).
        """
        rounds = _segment_into_rounds(records)
        if len(rounds) <= 1:
            return 100.0, {"transitions": 0, "pattern": "single_stage"}

        # Extract transition sequence (sequence of stages)
        transition_seq = [stage for stage, _ in rounds]

        transitions = len(rounds) - 1

        # Check regularity: count unexpected transitions
        # Compute how many "back-transitions" there are (e.g., D->P after P->D)
        back_transitions = 0
        prev_stage = transition_seq[0]
        seen_stages = {prev_stage}
        for stage in transition_seq[1:]:
            if stage != prev_stage:
                if stage in seen_stages and prev_stage != stage:
                    # Going back to a previously seen stage
                    back_transitions += 1
                seen_stages.add(stage)
                prev_stage = stage

        # Score based on regularity
        if back_transitions == 0:
            score = 100.0
        elif back_transitions <= 2:
            score = 80.0
        elif back_transitions <= 5:
            score = 60.0
        else:
            score = max(0.0, 60.0 * (1.0 - back_transitions / transitions)) if transitions > 0 else 0.0

        # Check round size regularity for decode rounds
        decode_sizes = [len(recs) for stage, recs in rounds if stage == "decode"]
        if len(decode_sizes) >= 2:
            size_cv, _ = _trimmed_cv([float(s) for s in decode_sizes])
            if size_cv > 0.2:
                score = max(0.0, score - 10.0)

        details = {
            "transitions": transitions,
            "back_transitions": back_transitions,
            "stage_sequence": transition_seq,
            "round_count": len(rounds),
        }
        return score, details

    # ── Legacy per-group methods (kept for diagnostics) ─────────────────

    def compute_time_consistency(self, records: List[LayerSummaryRecord]) -> Tuple[float, float, List[int]]:
        """Compute time consistency for a group (diagnostic).

        Returns (score, trimmed_cv, outlier_layer_indices).
        """
        times = [r.total_time_us for r in records]
        cv, outlier_positions = _trimmed_cv(times)
        outlier_indices = [records[i].layer_idx for i in outlier_positions]
        score = _score_time_cv(cv)
        return score, cv, outlier_indices

    def compute_kernel_count_consistency(self, records: List[LayerSummaryRecord]) -> Tuple[float, int, float, int]:
        """Compute kernel count consistency (diagnostic).

        Returns (score, mode_count, mode_pct, distinct_counts).
        Counts within ±2 of the mode are considered consistent.
        """
        counts = [r.kernel_count for r in records]
        counter = collections.Counter(counts)
        mode_count, _ = counter.most_common(1)[0]
        consistent_freq = sum(
            freq for val, freq in counter.items() if abs(val - mode_count) <= 2
        )
        mode_pct = 100.0 * consistent_freq / len(counts)
        distinct = len(counter)
        score = _score_mode_match(mode_pct)
        return score, mode_count, mode_pct, distinct

    def compute_pattern_consistency(self, patterns: List[LayerKernelPattern]) -> Tuple[float, float]:
        """Compute kernel pattern consistency (diagnostic).

        Uses fuzzy matching: two sequences are considered equivalent if their
        lengths differ by at most 2 and the shorter sequence is a subsequence
        of the longer one (i.e., the longer one has at most 2 extra kernels).

        Returns (score, mode_match_pct).
        """
        if not patterns:
            return 0.0, 0.0

        sequences = [p.short_names for p in patterns]
        n = len(sequences)

        # Find the mode sequence by exact match first
        counter = collections.Counter(sequences)
        mode_seq = counter.most_common(1)[0][0]

        # Count how many sequences are "close enough" to the mode:
        # length within ±2 AND the shorter is a subsequence of the longer
        match_count = 0
        for seq in sequences:
            if seq == mode_seq:
                match_count += 1
            elif abs(len(seq) - len(mode_seq)) <= 2 and _is_fuzzy_subsequence(seq, mode_seq):
                match_count += 1

        mode_pct = 100.0 * match_count / n
        score = _score_mode_match(mode_pct)
        return score, mode_pct

    # ── Per-group evaluation (diagnostic) ───────────────────────────────

    def _evaluate_group(
        self,
        key: GroupKey,
        records: List[LayerSummaryRecord],
        patterns: List[LayerKernelPattern],
    ) -> GroupMetrics:
        """Evaluate diagnostic metrics for one (stage, layer_type) group."""
        gm = GroupMetrics(key=key, layer_count=len(records))

        if len(records) < self.MIN_GROUP_SIZE:
            gm.skipped = True
            gm.skip_reason = f"< {self.MIN_GROUP_SIZE} layers"
            return gm

        # Prefill special handling: sub-group by kernel count if highly variable
        sub_records_list = [records]
        if key.stage == "prefill":
            kc_counter = collections.Counter(r.kernel_count for r in records)
            if len(kc_counter) > 3:
                sub_groups: Dict[int, List[LayerSummaryRecord]] = {}
                for r in records:
                    sub_groups.setdefault(r.kernel_count, []).append(r)
                sub_records_list = [sg for sg in sub_groups.values() if len(sg) >= self.MIN_GROUP_SIZE]
                if not sub_records_list:
                    sub_records_list = [records]
                gm.notes.append(f"Prefill sub-grouped by kernel count ({len(kc_counter)} distinct counts)")

        # Time consistency (diagnostic)
        total_weight = sum(len(sr) for sr in sub_records_list)
        agg_time_score = 0.0
        agg_time_cv = 0.0
        all_outlier_indices = []

        for sub_records in sub_records_list:
            w = len(sub_records) / total_weight
            score, cv, outliers = self.compute_time_consistency(sub_records)
            agg_time_score += w * score
            agg_time_cv += w * cv
            all_outlier_indices.extend(outliers)

        gm.time_score = agg_time_score
        gm.time_cv = agg_time_cv
        gm.outlier_indices = all_outlier_indices
        gm.outlier_count = len(all_outlier_indices)
        gm.outlier_pct = 100.0 * gm.outlier_count / len(records) if records else 0.0

        # Kernel count
        kc_score, kc_mode, kc_pct, kc_distinct = self.compute_kernel_count_consistency(records)
        gm.kernel_count_score = kc_score
        gm.kernel_count_mode = kc_mode
        gm.kernel_count_mode_pct = kc_pct
        gm.kernel_count_distinct = kc_distinct

        # Pattern
        if patterns:
            gm.pattern_available = True
            p_score, p_pct = self.compute_pattern_consistency(patterns)
            gm.pattern_score = p_score
            gm.pattern_mode_pct = p_pct
        else:
            gm.pattern_available = False

        # Outlier penalty (diagnostic)
        gm.outlier_score = _score_outlier_pct(gm.outlier_pct)

        # Composite (diagnostic, not used in overall scoring)
        w_time = 10.0
        w_kc = 50.0
        w_pat = 30.0 if gm.pattern_available else 0.0
        w_out = 10.0
        w_total = w_time + w_kc + w_pat + w_out
        if not gm.pattern_available:
            w_kc += 30.0
            w_total = w_time + w_kc + w_out

        gm.composite_score = (
            w_time * gm.time_score
            + w_kc * gm.kernel_count_score
            + w_pat * gm.pattern_score
            + w_out * gm.outlier_score
        ) / w_total

        return gm

    # ── Main evaluate ────────────────────────────────────────────────────

    def evaluate(self, filepath: str) -> EvaluationReport:
        """Evaluate an Excel file and return a full report."""
        overall, records, patterns = self.load_from_excel(filepath)
        return self._build_report(filepath, overall, records, patterns)

    def evaluate_from_result(self, result, filepath: str = "<AnalysisResult>") -> EvaluationReport:
        """Evaluate an AnalysisResult object and return a full report."""
        overall, records, patterns = self.load_from_analysis_result(result)
        return self._build_report(filepath, overall, records, patterns)

    def _build_report(
        self,
        filepath: str,
        overall: OverallStats,
        records: List[LayerSummaryRecord],
        patterns: List[LayerKernelPattern],
    ) -> EvaluationReport:
        """Build evaluation report from parsed data."""
        # ── Structural Rules ────────────────────────────────────────────
        sm = StructuralMetrics()

        sm.s1_score, sm.s1_details = self.compute_prefill_ordering(records)
        sm.s2_score, sm.s2_details = self.compute_architecture_signature(records)
        sm.s3_score, sm.s3_details = self.compute_round_consistency(records)
        sm.s4_score, sm.s4_details = self.compute_type_sequence_repeatability(records)

        total_weight = (
            W_S1_PREFILL_ORDERING + W_S2_ARCHITECTURE_SIG + W_S3_ROUND_CONSISTENCY
            + W_S4_TYPE_SEQUENCE
        )
        weighted_score = (
            W_S1_PREFILL_ORDERING * sm.s1_score
            + W_S2_ARCHITECTURE_SIG * sm.s2_score
            + W_S3_ROUND_CONSISTENCY * sm.s3_score
            + W_S4_TYPE_SEQUENCE * sm.s4_score
        )
        overall_score = weighted_score / total_weight

        # ── Per-group diagnostics ───────────────────────────────────────
        groups = self.group_layers(records, patterns)
        group_metrics_list = []
        all_outliers = []

        for key in sorted(groups.keys()):
            grp_records, grp_patterns = groups[key]
            gm = self._evaluate_group(key, grp_records, grp_patterns)
            group_metrics_list.append(gm)
            all_outliers.extend(gm.outlier_indices)

        return EvaluationReport(
            filepath=filepath,
            total_layers=len(records),
            group_count=len(groups),
            structural_metrics=sm,
            structural_score=round(overall_score, 1),
            group_metrics=group_metrics_list,
            overall_score=round(overall_score, 1),
            grade=_grade(overall_score),
            passed=overall_score >= self.threshold,
            threshold=self.threshold,
            all_outlier_indices=sorted(set(all_outliers)),
        )

    # ── Output formatting ────────────────────────────────────────────────

    def format_terminal_report(self, report: EvaluationReport) -> str:
        """Format report for terminal display."""
        lines = []
        sep = "=" * 70
        pass_fail = "PASS" if report.passed else "FAIL"

        lines.append("")
        lines.append(sep)
        lines.append("  Trace Parsing Quality Report")
        lines.append(f"  File: {report.filepath}")
        lines.append(f"  Layers: {report.total_layers} total, {report.group_count} groups")
        lines.append(f"  Overall Score: {report.overall_score} / 100 ({report.grade}) — {pass_fail}")
        lines.append(sep)

        # ── Structural Rules ─────────────────────────────────────────
        lines.append("")
        sm = report.structural_metrics

        # S1
        s1d = sm.s1_details
        lines.append(f"  S1 Prefill Ordering:    {sm.s1_score:.0f} ({_grade(sm.s1_score)})  violations={s1d.get('violations', 0)}")

        # S2
        s2d = sm.s2_details
        pat_str = str(s2d.get("detected_pattern", []))
        if len(pat_str) > 40:
            pat_str = pat_str[:37] + "..."
        lines.append(f"  S2 Architecture Sig:    {sm.s2_score:.0f} ({_grade(sm.s2_score)})  pattern={pat_str}")

        # S3
        s3d = sm.s3_details
        lines.append(
            f"  S3 Round Consistency:   {sm.s3_score:.0f} ({_grade(sm.s3_score)})  "
            f"decode_rounds={s3d.get('decode_rounds', 0)}, mode_size={s3d.get('mode_size', 0)}, "
            f"match={s3d.get('match_pct', 0)}%"
        )

        # S4
        s4d = sm.s4_details
        lines.append(
            f"  S4 Type Sequence:       {sm.s4_score:.0f} ({_grade(sm.s4_score)})  "
            f"rounds={s4d.get('rounds', 0)}, match={s4d.get('match_pct', 0)}%"
        )

        # ── Per-group diagnostics ────────────────────────────────────
        lines.append("")
        lines.append("  ── Group Diagnostics ──")

        for gm in report.group_metrics:
            lines.append("")
            if gm.skipped:
                lines.append(f"  --- {gm.key}: {gm.layer_count} layers — SKIPPED ({gm.skip_reason}) ---")
                continue

            lines.append(
                f"  --- {gm.key}: {gm.layer_count} layers — Diagnostic: {gm.composite_score:.1f} ---"
            )
            lines.append(
                f"    Time CV: {gm.time_cv:.3f}, {gm.outlier_count} outliers ({gm.outlier_pct:.1f}%)"
            )
            lines.append(
                f"    Kernel Count: mode={gm.kernel_count_mode} ({gm.kernel_count_mode_pct:.1f}%), "
                f"{gm.kernel_count_distinct} distinct"
            )
            if gm.pattern_available:
                lines.append(f"    Kernel Pattern: {gm.pattern_mode_pct:.1f}% match")
            for note in gm.notes:
                lines.append(f"    Note: {note}")

            # Sub-score breakdown table
            w_time = 10.0
            w_kc = 50.0
            w_pat = 30.0 if gm.pattern_available else 0.0
            w_out = 10.0
            if not gm.pattern_available:
                w_kc += 30.0
            lines.append(f"    ┌─────────────────┬────────┬────────┬───────┐")
            lines.append(f"    │ Metric          │ Score  │ Weight │ Grade │")
            lines.append(f"    ├─────────────────┼────────┼────────┼───────┤")
            lines.append(f"    │ Time CV         │ {gm.time_score:5.1f}  │ {w_time:5.0f}% │   {_grade(gm.time_score)}   │")
            lines.append(f"    │ Kernel Count    │ {gm.kernel_count_score:5.1f}  │ {w_kc:5.0f}% │   {_grade(gm.kernel_count_score)}   │")
            if gm.pattern_available:
                lines.append(f"    │ Kernel Pattern  │ {gm.pattern_score:5.1f}  │ {w_pat:5.0f}% │   {_grade(gm.pattern_score)}   │")
            lines.append(f"    │ Outlier Penalty │ {gm.outlier_score:5.1f}  │ {w_out:5.0f}% │   {_grade(gm.outlier_score)}   │")
            lines.append(f"    ├─────────────────┼────────┼────────┼───────┤")
            lines.append(f"    │ Composite       │ {gm.composite_score:5.1f}  │  100%  │   {_grade(gm.composite_score)}   │")
            lines.append(f"    └─────────────────┴────────┴────────┴───────┘")

        # Outlier list
        if report.all_outlier_indices:
            lines.append("")
            displayed = report.all_outlier_indices[:20]
            suffix = ", ..." if len(report.all_outlier_indices) > 20 else ""
            lines.append(f"  Outlier layers ({len(report.all_outlier_indices)} total): {displayed}{suffix}")

        lines.append("")
        return "\n".join(lines)

    def format_json_report(self, report: EvaluationReport) -> str:
        """Format report as JSON."""
        sm = report.structural_metrics
        data = {
            "filepath": report.filepath,
            "total_layers": report.total_layers,
            "group_count": report.group_count,
            "overall_score": report.overall_score,
            "grade": report.grade,
            "passed": report.passed,
            "threshold": report.threshold,
            "structural_rules": {
                "score": report.structural_score,
                "s1_prefill_ordering": {"score": round(sm.s1_score, 1), **sm.s1_details},
                "s2_architecture_sig": {"score": round(sm.s2_score, 1), **sm.s2_details},
                "s3_round_consistency": {"score": round(sm.s3_score, 1), **sm.s3_details},
                "s4_type_sequence": {"score": round(sm.s4_score, 1), **sm.s4_details},
            },
            "groups": [],
            "outlier_layer_count": len(report.all_outlier_indices),
            "outlier_layers": report.all_outlier_indices[:50],
        }

        for gm in report.group_metrics:
            g = {
                "stage": gm.key.stage,
                "layer_type": gm.key.layer_type,
                "layer_count": gm.layer_count,
                "skipped": gm.skipped,
            }
            if gm.skipped:
                g["skip_reason"] = gm.skip_reason
            else:
                w_kc = 80.0 if not gm.pattern_available else 50.0
                g.update({
                    "diagnostic_score": round(gm.composite_score, 1),
                    "time_consistency": {
                        "trimmed_cv": round(gm.time_cv, 4),
                        "score": round(gm.time_score, 1),
                        "weight": 10.0,
                        "outlier_count": gm.outlier_count,
                        "outlier_pct": round(gm.outlier_pct, 1),
                    },
                    "kernel_count": {
                        "mode": gm.kernel_count_mode,
                        "mode_pct": round(gm.kernel_count_mode_pct, 1),
                        "distinct_counts": gm.kernel_count_distinct,
                        "score": round(gm.kernel_count_score, 1),
                        "weight": w_kc,
                    },
                    "kernel_pattern": {
                        "available": gm.pattern_available,
                        "mode_pct": round(gm.pattern_mode_pct, 1),
                        "score": round(gm.pattern_score, 1) if gm.pattern_available else None,
                        "weight": 30.0 if gm.pattern_available else 0.0,
                    },
                    "outlier_penalty": {
                        "score": round(gm.outlier_score, 1),
                        "weight": 10.0,
                    },
                })
                if gm.notes:
                    g["notes"] = gm.notes

            data["groups"].append(g)

        return json.dumps(data, indent=2)


# ── Helpers ──────────────────────────────────────────────────────────────────

def _is_fuzzy_subsequence(seq_a: Tuple[str, ...], seq_b: Tuple[str, ...]) -> bool:
    """Check if the shorter sequence is a subsequence of the longer one.

    Used for fuzzy pattern matching: if two kernel sequences differ only by
    a few extra kernels (e.g., an extra ALLREDUCE), the shorter should appear
    as a subsequence within the longer.
    """
    short, long = (seq_a, seq_b) if len(seq_a) <= len(seq_b) else (seq_b, seq_a)
    it = iter(long)
    return all(item in it for item in short)


def _run_length_encode(sequence: List[str]) -> List[Tuple[int, str]]:
    """Run-length encode a sequence of strings.

    Returns [(count, value), ...].
    """
    if not sequence:
        return []

    result = []
    current = sequence[0]
    count = 1

    for item in sequence[1:]:
        if item == current:
            count += 1
        else:
            result.append((count, current))
            current = item
            count = 1

    result.append((count, current))
    return result


# ── CLI ──────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description="Evaluate trace parsing quality from analysis Excel files.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""\
Examples:
  python evaluate_parsing.py analysis.xlsx
  python evaluate_parsing.py analysis.xlsx --json
  python evaluate_parsing.py analysis.xlsx --threshold 85
""",
    )
    parser.add_argument("filepath", help="Path to analysis Excel file")
    parser.add_argument("--json", action="store_true", help="Output JSON instead of terminal report")
    parser.add_argument("--threshold", type=float, default=70.0, help="Pass/fail threshold (default: 70)")
    args = parser.parse_args()

    evaluator = ParsingEvaluator(threshold=args.threshold)
    report = evaluator.evaluate(args.filepath)

    if args.json:
        print(evaluator.format_json_report(report))
    else:
        print(evaluator.format_terminal_report(report))

    sys.exit(0 if report.passed else 1)


if __name__ == "__main__":
    main()
